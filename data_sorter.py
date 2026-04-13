from flask import Flask, send_from_directory, request, send_file, jsonify, redirect, session, make_response
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
import tempfile
import os
import time
import json
from typing import List, Dict, Any
from urllib.parse import quote

from title2vec_service import (
    process_titles,
    map_titles_to_families_and_seniority,
    map_titles_to_geography_and_country,
    _resolve_allowed_families  # internal helper
)
from gemini_context import fetch_gemini_project_context, fetch_gemini_skillset, fetch_gemini_project_date, infer_company_sector
from name_origin import (
    batch_detect_origins,
    origin_key_to_geo
)

# Load LLM provider config helper (for dynamic model selection)
try:
    from webbridge import _load_llm_provider_config as _ds_load_llm_cfg
except ImportError:
    _ds_load_llm_cfg = None

# Added imports for authentication/password handling
from werkzeug.security import check_password_hash

# Structured activity logger (shared with webbridge)
try:
    from app_logger import log_error as _log_error_ds, log_approval as _log_approval_ds
    _DS_LOGGER_AVAILABLE = True
except ImportError:
    _DS_LOGGER_AVAILABLE = False
    def _log_error_ds(**_kw): pass
    def _log_approval_ds(**_kw): pass


app = Flask(__name__, static_folder='static')

# Secret key for Flask session (use a secure value in production)
# If mounted by webbridge, this may be overwritten to ensure session compatibility
app.secret_key = os.getenv("FLASK_SECRET_KEY", "change-me-in-production")

# Register admin blueprint (optional). If admin_titles.py is present and importable,
# this will register the admin UI endpoints. If the module is missing or raises on import,
# we log the error and continue so the main app still starts.
try:
    from admin_titles import admin_bp
    app.register_blueprint(admin_bp)
    print("[Admin] admin_titles blueprint registered")
except Exception as _e:
    # Do not fail app startup if admin module is absent or erroneous.
    print("[Admin] admin_titles blueprint not registered:", _e)

# ========= Authentication gate (server-side) =========
def _get_username_from_request():
    uname = None
    try:
        uname = request.cookies.get('username')
    except Exception:
        uname = None
    if not uname:
        uname = request.args.get('username') or request.headers.get('X-Username')
    if not uname:
        try:
            j = request.get_json(silent=True) or {}
            if isinstance(j, dict):
                uname = j.get('username') or j.get('user') or j.get('user_name')
        except Exception:
            uname = None
    if uname:
        return str(uname).strip()
    return None

# ===== Affected section: user fetch + login =====
def _fetch_user_by_username(username: str):
    """
    Delegate to common_auth.fetch_user_by_username for consistent behavior across apps.
    """
    if not username:
        return None
    try:
        # Attempt direct import; if module is available on parent directory, ensure it's on sys.path
        try:
            import common_auth
        except Exception:
            import sys
            PARENT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
            if PARENT not in sys.path:
                sys.path.insert(0, PARENT)
            import common_auth
        return common_auth.fetch_user_by_username(username)
    except Exception as e:
        print("[Auth delegate fetch error]", e)
        return None

def _password_matches(stored: str, supplied: str) -> bool:
    """
    Delegate to common_auth.password_matches to support consistent password verification.
    """
    try:
        try:
            import common_auth
        except Exception:
            import sys
            PARENT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
            if PARENT not in sys.path:
                sys.path.insert(0, PARENT)
            import common_auth
        return common_auth.password_matches(stored, supplied)
    except Exception as e:
        print("[Auth delegate password check error]", e)
        return False

@app.route('/login', methods=['POST'])
def login_route():
    """
    Authenticate against login table. Expects JSON { username, password }.
    Uses common_auth.create_session_for_user to establish server session and cookies.
    """
    try:
        data = request.get_json(force=True)
    except Exception:
        return jsonify({"ok": False, "error": "Invalid JSON"}), 400
    username = (data.get('username') or "").strip()
    password = data.get('password') or ""
    if not username or not password:
        return jsonify({"ok": False, "error": "Missing credentials"}), 400

    user = _fetch_user_by_username(username)
    if not user:
        print("[Auth] Username not found:", username)
        return jsonify({"ok": False, "error": "Invalid username or password"}), 401

    stored = user.get('password') or ""
    if not _password_matches(stored, password):
        print("[Auth] Password verification failed for user:", username)
        return jsonify({"ok": False, "error": "Invalid username or password"}), 401

    # Successful authentication: create response and delegate session/cookie creation to common_auth
    resp = make_response(jsonify({
        "ok": True,
        "userid": user.get('id') or user.get('username'),
        "full_name": user.get('full_name') or "",
        "role_tag": user.get('role_tag') or ""
    }))
    secure_flag = not app.debug
    # Use common_auth to populate session and set cookies consistently
    try:
        try:
            import common_auth
        except Exception:
            import sys
            PARENT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
            if PARENT not in sys.path:
                sys.path.insert(0, PARENT)
            import common_auth
        # create_session_for_user will populate flask session and set username/userid cookies
        resp = common_auth.create_session_for_user(
            session,
            resp,
            user,
            app_debug=app.debug,
            cookie_samesite=os.getenv("COOKIE_SAMESITE", "Lax")
        )
    except Exception as e:
        # Fallback: set cookies similarly if common_auth fails for any reason
        print("[Auth] common_auth.create_session_for_user failed:", e)
        try:
            session['userid'] = user.get('id') or user.get('username')
            session['username'] = user['username']
            session['full_name'] = user.get('full_name') or ""
            session['role_tag'] = user.get('role_tag') or ""
            resp.set_cookie('username', user['username'], max_age=2592000, path='/', httponly=True, secure=secure_flag, samesite='Lax')
            resp.set_cookie('userid', str(session['userid']), max_age=2592000, path='/', httponly=True, secure=secure_flag, samesite='Lax')
        except Exception:
            pass

    print(f"[Auth] Login success username={username} userid={session.get('userid')}")
    return resp
# ===== End affected section =====

@app.route('/logout', methods=['POST', 'GET'])
def logout_route():
    try:
        session.clear()
    except Exception:
        pass
    resp = make_response(jsonify({"ok": True, "message": "Logged out"}))
    resp.set_cookie('username', '', max_age=0, path='/')
    resp.set_cookie('userid', '', max_age=0, path='/')
    return resp

@app.before_request
def require_login():
    if request.method == 'OPTIONS':
        return None

    path = request.path or '/'

    # Special case: Allow data_sorter.html only when embedded (iframe) via explicit signal
    # or when browser indicates request destination is an iframe.
    # Accepted signals:
    #  - Query parameter: ?embed=1
    #  - Header: X-Embed: true | 1
    #  - Browser Sec-Fetch-Dest header: 'iframe' (modern browsers)
    if path in ('/', '/data_sorter.html'):
        embed_param = (request.args.get('embed') or '').strip()
        embed_header = (request.headers.get('X-Embed') or '').strip().lower()
        sec_fetch = (request.headers.get('Sec-Fetch-Dest') or '').strip().lower()
        # Allow unauthenticated access if any embed indicator present
        if embed_param == '1' or embed_header in ('1', 'true') or sec_fetch == 'iframe':
            return None

    allow_prefixes = (
        '/static/',
        '/login',
        '/register',
        '/favicon.ico',
        '/download',
        '/user/resolve',
        '/user/update_role_tag',
        '/translate',
        '/translate_company',
        '/sector_suggest',
        '/suggest',
        '/preview_target',
        '/.well-known/'
    )
    for p in allow_prefixes:
        if path.startswith(p):
            return None
    if path.startswith('/static/'):
        return None
    try:
        sid = session.get('userid')
        sname = session.get('username')
        if sid and sname:
            request.environ['REMOTE_USERNAME'] = sname
            request.environ['REMOTE_USERID'] = sid
            return None
    except Exception:
        pass

    # Affected change: do NOT auto-create sessions from arbitrary query params/headers/JSON.
    # Use common_auth.restore_session_from_cookie to restore session from username cookie when possible.
    # For local development (app.debug) allow ?username auto-login only when request originates from localhost.
    user = None
    try:
        # Attempt to import common_auth (try direct import, else add parent folder to sys.path)
        try:
            import common_auth
        except Exception:
            import sys
            PARENT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
            if PARENT not in sys.path:
                sys.path.insert(0, PARENT)
            import common_auth
        # Try to restore session from cookie using shared helper
        user = common_auth.restore_session_from_cookie(request, session)
    except Exception:
        user = None

    # Development convenience: allow ?username auto-login when app.debug and request from localhost
    if not user and app.debug:
        try:
            q_uname = request.args.get('username')
            if q_uname and request.remote_addr in ('127.0.0.1', '::1', 'localhost'):
                try:
                    # fetch via common helper to ensure consistent user dict format
                    user = common_auth.fetch_user_by_username(q_uname)
                    if user:
                        session['userid'] = user.get('id') or user.get('username')
                        session['username'] = user.get('username')
                        session['full_name'] = user.get('full_name') or ""
                        session['role_tag'] = user.get('role_tag') or ""
                except Exception:
                    user = None
        except Exception:
            user = None

    if user:
        try:
            request.environ['REMOTE_USERNAME'] = user.get('username')
            request.environ['REMOTE_USERID'] = session.get('userid')
        except Exception:
            pass
        return None

    accepts_html = False
    try:
        accepts_html = request.accept_mimetypes.accept_html
    except Exception:
        accepts_html = False
    if request.method == 'GET' and accepts_html:
        qs = ''
        try:
            qs = ('?' + request.query_string.decode()) if request.query_string else ''
        except Exception:
            qs = ''
        next_path = request.path + qs
        return redirect('/login.html?next=' + quote(next_path))
    return jsonify({"error": "Unauthorized", "message": "Authentication required"}), 401

QA_ORIGIN_CLUSTER_MIN = int(os.getenv("QA_ORIGIN_CLUSTER_MIN", "6"))
QA_ORIGIN_CLUSTER_RATIO = float(os.getenv("QA_ORIGIN_CLUSTER_RATIO", "0.55"))
MIN_ORIGIN_CONFIDENCE = float(os.getenv("MIN_ORIGIN_CONFIDENCE", "0.6"))
ROLE_VP_EMEA_REGION = os.getenv("ROLE_VP_EMEA_REGION", "Western Europe")
ROLE_VP_EMEA_COUNTRY = os.getenv("ROLE_VP_EMEA_COUNTRY", "")
QA_LIKE_TOKENS = ["qa", "test", "tester", "quality assurance"]
ORIGIN_DEBUG = os.getenv("ORIGIN_DEBUG", "0") == "1"

try:
    with open(os.path.join("static", "data_sorter.json"), encoding="utf-8") as _f:
        _GEO_MAP = json.load(_f).get("GeoCountries", {})
except Exception:
    _GEO_MAP = {}

def region_from_country(country: str) -> str:
    c = (country or "").strip()
    if not c:
        return ""
    for region, countries in _GEO_MAP.items():
        if c in countries:
            return region
    base = c.split(" (")[0].strip().lower()
    for region, countries in _GEO_MAP.items():
        for cc in countries:
            if base == cc.split(" (")[0].strip().lower():
                return region
    return ""

ORG_SYNONYMS = {"organisation", "organization", "organsation", "company"}

def find_existing_org_header(uploaded_columns: List[str]) -> str:
    for h in uploaded_columns:
        if isinstance(h, str) and h.strip().lower() in ORG_SYNONYMS:
            return h
    return ""

_SKILLSET_CACHE: Dict[tuple, List[str]] = {}

def detect_title_column(headers):
    if not headers: return None
    kws = ["title", "job", "position", "role", "jobtitle"]
    for i,h in enumerate(headers):
        if isinstance(h,str) and any(k in h.lower() for k in kws):
            return i
    return None

def detect_name_column(headers):
    if not headers: return None
    kws = ["name","candidate"]
    for i,h in enumerate(headers):
        if isinstance(h,str) and any(k in h.lower() for k in kws):
            return i
    return None

def normalize_rows(uploaded_columns: List[str], rows: List[Any]) -> List[List[Any]]:
    if not rows:
        return []
    first = rows[0]
    if not isinstance(first, dict):
        return rows
    def build_keymap(d):
        return {k.lower(): k for k in d.keys()}
    normalized = []
    for r in rows:
        if not isinstance(r, dict):
            normalized.append(r)
            continue
        keymap = build_keymap(r)
        row_arr = []
        for h in uploaded_columns:
            if h in r:
                row_arr.append(r.get(h, ""))
            else:
                lk = h.lower() if isinstance(h, str) else ""
                mapped = keymap.get(lk)
                if mapped is not None:
                    row_arr.append(r.get(mapped, ""))
                else:
                    row_arr.append(r.get(h.lower(), "") if isinstance(h, str) else "")
        normalized.append(row_arr)
    return normalized

def extract_titles(uploaded_columns, rows):
    if not uploaded_columns:
        return ["" for _ in rows]
    if rows and isinstance(rows[0], dict):
        rows = normalize_rows(uploaded_columns, rows)
    idx = detect_title_column(uploaded_columns)
    if idx is None:
        return ["" for _ in rows]
    out=[]
    for r in rows:
        v = r[idx] if idx < len(r) else ""
        out.append(v.strip() if isinstance(v,str) else "")
    return out

def extract_names(uploaded_columns, rows):
    if not uploaded_columns:
        return ["" for _ in rows]
    if rows and isinstance(rows[0], dict):
        rows = normalize_rows(uploaded_columns, rows)
    idx = detect_name_column(uploaded_columns)
    if idx is None:
        return ["" for _ in rows]
    out=[]
    for r in rows:
        v = r[idx] if idx < len(r) else ""
        out.append(v.strip() if isinstance(v,str) else "")
    return out

def get_headers(data):
    uploaded_columns = data.get('uploaded_columns', [])
    extra=[]
    existing_org = find_existing_org_header(uploaded_columns)
    if data.get('company') and not existing_org:
        extra.append("company")
    if data.get('product') and "product" not in uploaded_columns:
        extra.append("product")
    if data.get('project_date') and "project_date" not in uploaded_columns:
        extra.append("project_date")
    if data.get('sector') or data.get('sector_other'):
        extra.append("sector")
    if data.get('jobfamily') or data.get('jobfamily_other'):
        extra.append("jobfamily")
    if data.get('role_tag'):
        extra.append("role_tag")
    if data.get('geographics') or data.get('geo_other') or data.get('infer_location'):
        extra.append("geographic")
    if data.get('country') or data.get('infer_location'):
        extra.append("country")
    contact_options=data.get('contact_options',[])
    contact_other=(data.get('contact_other') or "").strip()
    contact_cols=[c for c in contact_options if c!="Other"]
    if contact_other:
        contact_cols.append(contact_other)
    elif "Other" in contact_options:
        contact_cols.append("Other")
    for c in contact_cols:
        if c not in uploaded_columns:
            extra.append(c)
    if data.get('seniority') or data.get('seniority_other') or data.get('infer_seniority'):
        extra.append("seniority")
    if "skillset" not in uploaded_columns:
        extra.append("skillset")
    if data.get('sourcingstatus') or data.get('sourcing_status_other'):
        extra.append("sourcingstatus")
    header = uploaded_columns + [c for c in extra if c not in uploaded_columns]
    if "project_date" in header:
        if "product" in header:
            pt_index = header.index("product")
            pd_index = header.index("project_date")
            if pd_index != pt_index + 1:
                header.pop(pd_index)
                header.insert(pt_index + 1, "project_date")
    return header

def detect_role_based_overrides(titles: List[str]) -> Dict[int, Dict[str, Any]]:
    overrides={}
    for i,t in enumerate(titles):
        tl=t.lower()
        if "emea" in tl and ("vp" in tl or "vice president" in tl):
            overrides[i]={
                "region": ROLE_VP_EMEA_REGION,
                "country": ROLE_VP_EMEA_COUNTRY,
                "annotation": f"Role override: VP EMEA -> {ROLE_VP_EMEA_REGION}"
            }
    return overrides

def build_per_row_origin_overrides(
    titles: List[str],
    names: List[str],
    default_country: str
) -> Dict[int, Dict[str, Any]]:
    overrides={}
    if not titles:
        return overrides
    row_origins = batch_detect_origins(names)
    title_groups: Dict[str,List[int]] = {}
    for i,t in enumerate(titles):
        if t:
            title_groups.setdefault(t.lower(), []).append(i)
    for t_lower, idxs in title_groups.items():
        if not any(tok in t_lower for tok in QA_LIKE_TOKENS):
            continue
        if len(idxs) < QA_ORIGIN_CLUSTER_MIN:
            continue
        foreign_counts: Dict[str,int] = {}
        jp_present=False
        for i in idxs:
            o=row_origins[i]
            if not o: continue
            ok=o["origin_key"]
            if ok=="japanese":
                jp_present=True
            else:
                foreign_counts[ok]=foreign_counts.get(ok,0)+1
        if jp_present:
            continue
        total=len(idxs)
        for ok,cnt in foreign_counts.items():
            ratio=cnt/total
            if cnt >= QA_ORIGIN_CLUSTER_MIN and ratio >= QA_ORIGIN_CLUSTER_RATIO:
                geo=origin_key_to_geo(ok)
                if not geo: continue
                for i in idxs:
                    o=row_origins[i]
                    if not o or o["origin_key"]!=ok: continue
                    if o.get("confidence",0) < MIN_ORIGIN_CONFIDENCE: continue
                    if default_country.lower()=="japan" or geo["country"].lower()!=default_country.lower():
                        overrides[i]={
                            "country": geo["country"],
                            "region": geo["region"],
                            "annotation": f"Foreign QA partial cluster {ok} {cnt}/{total} (title '{t_lower}')"
                        }
                if ORIGIN_DEBUG:
                    print(f"[OriginOverride] title='{t_lower}' origin={ok} size={cnt}/{total}")
    return overrides

def generate_excel(data):
    generate_skillset = bool(data.get('generate_skillset', True)) and os.getenv("DISABLE_SKILLSET", "0") != "1"
    wb=openpyxl.Workbook()
    ws=wb.active
    ws.title="Selections"
    header=get_headers(data)
    ws.append(header)
    rows=data.get('rows',[])
    uploaded_cols=data.get('uploaded_columns',[])
    rows = normalize_rows(uploaded_cols, rows)
    sector = data.get("sector") or ""
    sector_norm = sector.lower().strip()
    titles=extract_titles(uploaded_cols, rows)
    need_family="jobfamily" in header
    need_seniority="seniority" in header
    inferred_families,inferred_seniorities=([],[])
    if need_family or need_seniority:
        inferred_families,inferred_seniorities=map_titles_to_families_and_seniority(titles, sector=sector_norm)
    infer_location=bool(data.get('infer_location'))
    default_region=data.get("default_region") or ""
    default_country=data.get("default_country") or ""
    need_geo=("geographic" in header or "country" in header) and infer_location
    inferred_regions,inferred_countries=([],[])
    if need_geo:
        inferred_regions,inferred_countries=map_titles_to_geography_and_country(
            titles, default_region=default_region, default_country=default_country
        )
    row_override_map: Dict[int, Dict[str,str]]={}
    for override_list_name in ["origin_overrides","role_overrides"]:
        for ov in data.get(override_list_name, []):
            ri=ov.get("row_index")
            if isinstance(ri,int):
                row_override_map[ri]={
                    "country": ov.get("country",""),
                    "region": ov.get("region","")
                }
    sector_val = sector
    if sector_val == "Other":
        sector_val = data.get("sector_other") or ""
    contact_options=data.get('contact_options',[])
    contact_other=(data.get('contact_other') or "").strip()
    contact_cols=[c for c in contact_options if c!="Other"]
    if contact_other:
        if contact_other not in contact_cols:
            contact_cols.append(contact_other)
    elif "Other" in contact_options:
        contact_cols.append("Other")

    # Ensure contact_indices is initialized before use
    contact_indices = {}
    for c in contact_cols:
        if c in header:
            contact_indices[c]=header.index(c)
    seniorities_user=data.get('seniority',[])
    if seniorities_user:
        seniorities_user=[s for s in seniorities_user if s!="Other"]
    job_families_selected=data.get('jobfamily',[])
    company_name=(data.get('company') or "").strip()
    product_name=""
    if not bool(data.get('use_process_table')):
        product_name=(data.get('product') or "").strip()
    project_date=(data.get('project_date') or "").strip()
    existing_org_header = find_existing_org_header(uploaded_cols)
    if not existing_org_header and "company" in header:
        existing_org_header = "company"
    col_indices={h:i for i,h in enumerate(header)}
    for i,original_row in enumerate(rows):
        row=list(original_row)
        def ensure(idx):
            while len(row)<=idx:
                row.append("")
        if existing_org_header and existing_org_header in col_indices:
            oi = col_indices[existing_org_header]
            ensure(oi)
            if not row[oi] and company_name:
                row[oi]=company_name
        if "product" in col_indices:
            ensure(col_indices["product"])
            if not row[col_indices["product"]]:
                row[col_indices["product"]]=product_name
        if "project_date" in col_indices:
            ensure(col_indices["project_date"])
            if not row[col_indices["project_date"]] and project_date:
                row[col_indices["project_date"]] = project_date
        if "sector" in col_indices:
            ensure(col_indices["sector"])
            if not row[col_indices["sector"]]:
                row[col_indices["sector"]]=sector_val
        if "jobfamily" in col_indices:
            ensure(col_indices["jobfamily"])
            if (not row[col_indices["jobfamily"]]) and i < len(inferred_families):
                row[col_indices["jobfamily"]] = inferred_families[i]
        if "seniority" in col_indices:
            ensure(col_indices["seniority"])
            if (not row[col_indices["seniority"]]) and i < len(inferred_seniorities):
                row[col_indices["seniority"]] = inferred_seniorities[i]
        region_val=default_region
        country_val=default_country
        if need_geo:
            if i < len(inferred_regions) and inferred_regions[i]:
                region_val=inferred_regions[i]
            if i < len(inferred_countries) and inferred_countries[i]:
                country_val=inferred_countries[i]
        if i in row_override_map:
            ov=row_override_map[i]
            if ov.get("region"):
                region_val=ov["region"]
            if ov.get("country") or ov.get("country")=="":
                country_val=ov["country"]
        if not region_val and not country_val:
            region_val=default_region
            country_val=default_country
        if "country" in col_indices:
            ensure(col_indices["country"])
            if not row[col_indices["country"]]:
                row[col_indices["country"]] = country_val
        if "geographic" in col_indices:
            ensure(col_indices["geographic"])
            if not row[col_indices["geographic"]]:
                derived = region_val or ""
                if not derived:
                    existing_country = row[col_indices["country"]] if "country" in col_indices and col_indices["country"] < len(row) else ""
                    if existing_country:
                        derived = region_from_country(existing_country)
                if derived:
                    row[col_indices["geographic"]] = derived
        if generate_skillset and "skillset" in col_indices:
            si = col_indices["skillset"]
            ensure(si)
            if not row[si]:
                job_title_val = titles[i] if i < len(titles) else ""
                job_family_val = row[col_indices["jobfamily"]] if "jobfamily" in col_indices and col_indices["jobfamily"] < len(row) else ""
                seniority_val = row[col_indices["seniority"]] if "seniority" in col_indices and col_indices["seniority"] < len(row) else ""
                country_cell_val = row[col_indices["country"]] if "country" in col_indices and col_indices["country"] < len(row) else ""
                organisation_val = row[col_indices[existing_org_header]] if existing_org_header and existing_org_header in col_indices and col_indices[existing_org_header] < len(row) else company_name
                cache_key = (
                    job_title_val.strip().lower(),
                    (organisation_val or "").strip().lower(),
                    "" if bool(data.get('use_process_table')) else product_name.strip().lower(),
                    country_cell_val.strip().lower(),
                    job_family_val.strip().lower(),
                    seniority_val.strip().lower()
                )
                skills = _SKILLSET_CACHE.get(cache_key)
                if skills is None:
                    skills = fetch_gemini_skillset(
                        jobtitle=job_title_val,
                        company=organisation_val,
                        project_title="" if bool(data.get('use_process_table')) else product_name,
                        country=country_cell_val,
                        job_family=job_family_val,
                        seniority=seniority_val
                    )
                    _SKILLSET_CACHE[cache_key] = skills
                if skills:
                    row[si] = "; ".join(skills)
        for c, idx_c in contact_indices.items():
            ensure(idx_c)
            if not row[idx_c]:
                row[idx_c]=""
        if "sourcingstatus" in col_indices:
            ensure(col_indices["sourcingstatus"])
            if not row[col_indices["sourcingstatus"]]:
                row[col_indices["sourcingstatus"]]=""
        row+=[""]*(len(header)-len(row))
        ws.append(row[:len(header)])
    data_rows_start=2
    data_rows_end=ws.max_row
    def add_dropdown(col_name, options):
        if not options or col_name not in col_indices:
            return
        col_number=col_indices[col_name]+1
        cleaned=[]
        seen=set()
        for o in options:
            if not o: continue
            c=str(o).replace(",", "").replace('"',"").replace("\n","").strip()
            if c and c not in seen:
                seen.add(c); cleaned.append(c)
        if not cleaned: return
        col_letter=openpyxl.utils.get_column_letter(col_number)
        cell_range=f"{col_letter}{data_rows_start}:{col_letter}{data_rows_end}"
        joined=",".join(cleaned)
        if len(joined)<=255:
            dv=DataValidation(type="list", formula1=f'"{joined}"', allow_blank=True)
            ws.add_data_validation(dv); dv.add(cell_range)
        else:
            list_sheet="Lists"
            if list_sheet not in wb.sheetnames:
                ls=wb.create_sheet(list_sheet)
                ls.sheet_state='hidden'
            else:
                ls=wb[list_sheet]
            write_col=1
            while ls.cell(row=1,column=write_col).value is not None:
                write_col+=1
            for i,val in enumerate(cleaned,1):
                ls.cell(row=i,column=write_col,value=val)
            letter=openpyxl.utils.get_column_letter(write_col)
            ref=f"{list_sheet}!${letter}$1:${letter}${len(cleaned)}"
            dv=DataValidation(type="list", formula1=f"={ref}", allow_blank=True)
            ws.add_data_validation(dv); dv.add(cell_range)
    if "sector" in col_indices and sector_val:
        add_dropdown("sector",[sector_val])
    if "jobfamily" in col_indices:
        allowed_sector_fams = _resolve_allowed_families(sector_norm) or []
        all_fams = list(dict.fromkeys(
            [f for f in inferred_families if f and f in allowed_sector_fams] +
            [f for f in job_families_selected if f in allowed_sector_fams]
        ))
        add_dropdown("jobfamily", all_fams)
    if "seniority" in col_indices:
        inferred_unique=[s for s in dict.fromkeys(inferred_seniorities) if s]
        base=list(dict.fromkeys((seniorities_user or []) + inferred_unique))
        if not base:
            base=["Junior","Mid","Senior","Lead","Manager","Director","Expert","Executive"]
        add_dropdown("seniority", base)
    if "sourcingstatus" in col_indices:
        add_dropdown("sourcingstatus", data.get('sourcingstatus', []))
    if "geographic" in col_indices:
        geos=set()
        for r in range(data_rows_start, ws.max_row+1):
            val=ws.cell(row=r, column=col_indices["geographic"]+1).value
            if val: geos.add(val)
        add_dropdown("geographic", sorted(geos))
    if "country" in col_indices:
        countries=set()
        for r in range(data_rows_start, ws.max_row+1):
            val=ws.cell(row=r, column=col_indices["country"]+1).value
            if val: countries.add(val)
        add_dropdown("country", sorted(countries))
    tmp=tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(tmp.name)
    tmp.close()
    return tmp.name

@app.route('/infer_job_families', methods=['POST'])
def infer_job_families_route():
    start=time.time()
    data=request.get_json(force=True)
    fetch_from_db = bool(data.get('use_process_table') or data.get('fetch_process') or data.get('from_process_table'))
    uploaded_columns = data.get('uploaded_columns',[])
    rows = data.get('rows',[])
    if not fetch_from_db and (not rows and (data.get('userid') or data.get('username'))):
        fetch_from_db = True
    if fetch_from_db:
        try:
            import psycopg2
            from psycopg2 import sql
            pg_host=os.getenv("PGHOST","localhost")
            pg_port=int(os.getenv("PGPORT","5432"))
            pg_user=os.getenv("PGUSER","postgres")
            pg_password=os.getenv("PGPASSWORD","") or "orlha"
            pg_db=os.getenv("PGDATABASE","candidate_db")
            conn=psycopg2.connect(host=pg_host, port=pg_port, user=pg_user, password=pg_password, dbname=pg_db)
            cur=conn.cursor()
            cur.execute("""
                SELECT column_name
                FROM information_schema.columns
                WHERE table_schema='public' AND table_name='process'
            """)
            cols = [r[0].lower() for r in cur.fetchall()]
            if not cols:
                cur.close(); conn.close()
                return jsonify({"error":"Process table not found or has no columns"}), 500
            preferred_title_col = None
            if 'jobtitle' in cols:
                preferred_title_col = 'jobtitle'
            elif 'role' in cols:
                preferred_title_col = 'role'
            else:
                for c in cols:
                    if 'title' in c or 'role' in c:
                        preferred_title_col = c
                        break
            if not preferred_title_col:
                cur.close(); conn.close()
                return jsonify({"error":"No suitable title/role column found in process table"}), 500
            candidate_cols = []
            if 'id' in cols and 'id' not in candidate_cols:
                candidate_cols.append('id')
            for c in ['name','company', preferred_title_col, 'country', 'linkedinurl', 'sector', 'username', 'userid', 'role_tag']:
                if c in cols and c not in candidate_cols:
                    candidate_cols.append(c)
            if not candidate_cols:
                cur.close(); conn.close()
                return jsonify({"error":"No usable columns present in process table"}), 500

            include_ctid = False
            if 'id' not in cols:
                include_ctid = True

            select_parts = []
            for c in candidate_cols:
                select_parts.append(sql.Identifier(c))
            if include_ctid:
                select_parts.append(sql.SQL("ctid::text AS _ctid"))

            select_fields = sql.SQL(', ').join(select_parts)
            base_query = sql.SQL("SELECT {fields} FROM process").format(fields=select_fields)

            where_clauses = []
            params = []
            payload_userid = data.get('userid')
            payload_username = data.get('username')
            sess_userid = session.get('userid')
            sess_username = session.get('username')
            userid = payload_userid if payload_userid is not None else sess_userid
            username = payload_username if payload_username is not None else sess_username
            if userid:
                where_clauses.append(sql.SQL("userid = %s"))
                params.append(str(userid))
            elif username:
                where_clauses.append(sql.SQL("username = %s"))
                params.append(username)
            limit = int(data.get('limit', 1000) or 1000)
            if where_clauses:
                where_sql = sql.SQL(" WHERE ") + sql.SQL(" AND ").join(where_clauses)
                query = base_query + where_sql + sql.SQL(" ORDER BY name NULLS LAST LIMIT %s")
                params.append(limit)
            else:
                query = base_query + sql.SQL(" ORDER BY name NULLS LAST LIMIT %s")
                params = [limit]
            cur.execute(query.as_string(conn), params)
            fetched = cur.fetchall()

            uploaded_columns = candidate_cols[:]
            if include_ctid:
                uploaded_columns.append('_ctid')

            rows = []
            for r in fetched:
                row_dict = {}
                for idx, col in enumerate(uploaded_columns):
                    row_dict[col] = r[idx]
                rows.append(row_dict)
            cur.close(); conn.close()
        except Exception as e:
            import traceback
            traceback.print_exc()
            return jsonify({"error": f"Process table fetch failed: {e}"}), 500
    # preserve original rows (fetched or provided) for server-side counts before normalization
    original_rows_for_count = rows
    rows = normalize_rows(uploaded_columns, rows)
    top_n=data.get('top_n',30)
    company=(data.get("company") or "").strip()
    product=""
    if not fetch_from_db:
        product=(data.get("product") or "").strip()
    sector=(data.get("sector") or "").strip()
    sector_norm = sector.lower()
    use_gemini=bool(data.get("infer_gemini"))
    gemini_model=(data.get("gemini_model") or "").strip() or None
    # Fall back to the globally configured Gemini model when no per-request model is given
    if not gemini_model and _ds_load_llm_cfg:
        try:
            _llm_cfg = _ds_load_llm_cfg()
            _gem_cfg = _llm_cfg.get("gemini", {})
            _cfgmodel = (_gem_cfg.get("model") or "").strip()
            if _cfgmodel:
                gemini_model = _cfgmodel
        except Exception:
            pass
    infer_location = bool(data.get('infer_location'))
    titles = extract_titles(uploaded_columns, rows)
    # --- New: compute rendering count (role_tag == 'Rendering') from original_rows_for_count when available ---
    try:
        rendering_count = 0
        target_role = 'rendering'
        if original_rows_for_count and len(original_rows_for_count):
            # rows may be list of dicts (from DB) or list of lists (from payload)
            first = original_rows_for_count[0]
            if isinstance(first, dict):
                for r in original_rows_for_count:
                    if not isinstance(r, dict):
                        continue
                    # check known keys
                    val = None
                    if 'role_tag' in r:
                        val = r.get('role_tag')
                    elif 'role' in r:
                        val = r.get('role')
                    else:
                        # case-insensitive search
                        for k,v in r.items():
                            try:
                                if str(k).strip().lower().replace(' ','').replace('_','') in ('roletag','role'):
                                    val = v
                                    break
                            except Exception:
                                continue
                    if val and str(val).strip().lower() == target_role:
                        rendering_count += 1
            else:
                # assume list-of-lists and uploaded_columns contains headers
                lower_headers = [str(h).lower() for h in uploaded_columns]
                # find role index
                role_idx = None
                for i,h in enumerate(lower_headers):
                    if h.strip() in ('role_tag','roletag','role tag') or h.strip().startswith('role'):
                        role_idx = i
                        break
                if role_idx is not None:
                    for r in original_rows_for_count:
                        if not isinstance(r, (list, tuple)):
                            continue
                        if role_idx < len(r):
                            val = r[role_idx]
                            if val and str(val).strip().lower() == target_role:
                                rendering_count += 1
                else:
                    # fallback: scan all cells for exact 'rendering'
                    for r in original_rows_for_count:
                        if not isinstance(r, (list, tuple)):
                            continue
                        matched = False
                        for c in r:
                            if c and str(c).strip().lower() == target_role:
                                rendering_count += 1
                                matched = True
                                break
                        if matched:
                            continue
        else:
            rendering_count = 0
    except Exception:
        rendering_count = 0
    # --- end rendering_count calculation ---

    if not any(titles):
        elapsed=time.time()-start
        return jsonify({
            "suggestions": [],
            "per_title_mapping": [],
            "message": "No recognizable job title column found.",
            "processing_seconds": elapsed,
            "geographic_suggestions": [],
            "country_suggestions": [],
            "project_context": {},
            "default_region": "",
            "default_country": "",
            "origin_overrides": [],
            "role_overrides": [],
            "location_annotations": [],
            "rendering_count": rendering_count
        })
    project_context=fetch_gemini_project_context(company, product, model_name=gemini_model) if use_gemini else {}
    default_region=project_context.get("primary_region") if project_context else None
    default_country=project_context.get("primary_country") if project_context else None
    result=process_titles(
        list(dict.fromkeys([t for t in titles if t])),
        top_n=top_n,
        default_region=default_region,
        default_country=default_country,
        sector=sector_norm
    )
    per_map=result.get("per_title_mapping",[])
    title_entry_map={e["title"].lower(): e for e in per_map}

    # === AFFECTED SECTION: enrich per_map entries with role_tag where possible ===
    # Purpose: ensure the frontend can synthesize CSVs (and compute counts) using role_tag present
    try:
        if uploaded_columns and rows and per_map:
            # Normalize header names to lower-case for index discovery
            lower_headers = [str(h).lower() for h in uploaded_columns]
            title_idx = detect_title_column(uploaded_columns)
            # find probable role_tag column
            role_idx = None
            for i, h in enumerate(lower_headers):
                if h in ('role_tag','roletag','role tag') or h == 'role' or h.startswith('role'):
                    role_idx = i
                    break
            company_idx = None
            for i, h in enumerate(lower_headers):
                if h == 'company':
                    company_idx = i
                    break

            # Build a mapping: normalized title -> counts of role_tag values (optionally by company)
            role_counts_by_title: Dict[str, Dict[str,int]] = {}
            for r in rows:
                if not isinstance(r, (list, tuple)):
                    continue
                # extract title and role_tag values safely
                try:
                    tval = ""
                    rval = ""
                    cval = ""
                    if title_idx is not None and title_idx < len(r):
                        tval = r[title_idx] or ""
                    if role_idx is not None and role_idx < len(r):
                        rval = r[role_idx] or ""
                    if company_idx is not None and company_idx < len(r):
                        cval = r[company_idx] or ""
                    if not tval:
                        continue
                    key = str(tval).strip().lower()
                    if not key:
                        continue
                    rv = str(rval).strip() if rval is not None else ""
                    if rv:
                        d = role_counts_by_title.setdefault(key, {})
                        d[rv] = d.get(rv, 0) + 1
                except Exception:
                    continue

            # Apply most-common role_tag to per_map entries when available
            for entry in per_map:
                try:
                    etitle = (entry.get("title") or "").strip().lower()
                    if not etitle:
                        continue
                    if etitle in role_counts_by_title:
                        counts = role_counts_by_title[etitle]
                        # pick the most frequent non-empty role tag
                        sorted_roles = sorted([ (k,v) for k,v in counts.items() if k ], key=lambda x: x[1], reverse=True)
                        if sorted_roles:
                            entry["role_tag"] = sorted_roles[0][0]
                        else:
                            entry["role_tag"] = entry.get("role_tag", "")
                    else:
                        # ensure the key exists (empty) so frontend sees the field
                        entry["role_tag"] = entry.get("role_tag", "")
                except Exception:
                    # safe continue
                    continue
    except Exception:
        pass
    # === END AFFECTED SECTION ===

    names=extract_names(uploaded_columns, rows)
    origin_overrides=build_per_row_origin_overrides(titles, names, default_country or "")
    role_overrides=detect_role_based_overrides(titles)
    for row_index,ov in origin_overrides.items():
        if row_index < 0 or row_index >= len(titles):
            continue
        t=titles[row_index].lower()
        entry=title_entry_map.get(t)
        if entry:
            if ov.get("country") or ov.get("country")=="":
                entry["country"]=ov["country"]
            if ov.get("region"):
                entry["geographic"]=ov["region"]
    for row_index,ov in role_overrides.items():
        if row_index < 0 or row_index >= len(titles):
            continue
        t=titles[row_index].lower()
        entry=title_entry_map.get(t)
        if entry:
            if ov.get("country") or ov.get("country")=="":
                entry["country"]=ov["country"]
            if ov.get("region"):
                entry["geographic"]=ov["region"]

    use_country_map = False
    try:
        if fetch_from_db and infer_location and uploaded_columns:
            lower_headers = [str(h).lower() for h in uploaded_columns]
            if 'country' in lower_headers:
                use_country_map = True
    except Exception:
        use_country_map = False

    if use_country_map:
        title_to_country_counts: Dict[str, Dict[str, int]] = {}
        title_idx = detect_title_column(uploaded_columns)
        country_idx = None
        for i,h in enumerate(uploaded_columns):
            try:
                if isinstance(h, str) and h.strip().lower() == 'country':
                    country_idx = i
                    break
            except Exception:
                continue
        for r in rows:
            if not isinstance(r, (list, tuple)):
                continue
            tval = ""
            cval = ""
            if title_idx is not None and title_idx < len(r):
                tval = r[title_idx] if isinstance(r[title_idx], str) else (str(r[title_idx]) if r[title_idx] is not None else "")
                tval = tval.strip()
            if country_idx is not None and country_idx < len(r):
                cval = r[country_idx] if isinstance(r[country_idx], str) else (str(r[country_idx]) if r[country_idx] is not None else "")
                cval = cval.strip()
            if not tval or not cval:
                continue
            tl = tval.lower()
            title_to_country_counts.setdefault(tl, {})
            title_to_country_counts[tl][cval] = title_to_country_counts[tl].get(cval, 0) + 1
        for title_lower, counts in title_to_country_counts.items():
            entry = title_entry_map.get(title_lower)
            if not entry:
                continue
            sorted_c = sorted(counts.items(), key=lambda x: x[1], reverse=True)
            if sorted_c:
                country_choice = sorted_c[0][0]
                entry["country"] = country_choice
                entry["geographic"] = region_from_country(country_choice) or entry.get("geographic", "")

    # Collect title -> sector mapping from the process table if present (ensure DB-written sector is preferred)
    try:
        if fetch_from_db and uploaded_columns:
            lower_headers = [str(h).lower() for h in uploaded_columns]
            title_idx = detect_title_column(uploaded_columns)
            sector_idx = None
            company_idx = None
            for i,h in enumerate(uploaded_columns):
                try:
                    if isinstance(h, str):
                        hl = h.strip().lower()
                        if hl == 'sector':
                            sector_idx = i
                        if hl == 'company':
                            company_idx = i
                except Exception:
                    continue

            # Only proceed if we have a title and sector column
            if title_idx is not None and sector_idx is not None:
                import re
                def _normalize_title_for_matching(t: str) -> str:
                    if not t:
                        return ""
                    s = str(t).lower().strip()
                    # remove parentheses content
                    s = re.sub(r'\(.*?\)', ' ', s)
                    # remove punctuation
                    s = re.sub(r'[^\w\s]', ' ', s)
                    # remove common seniority tokens to improve matching
                    s = re.sub(r'\b(?:senior|sr|sr\.|jr|jr\.|junior|lead|principal|manager|mgr|director|vp|head|staff|associate)\b', ' ', s, flags=re.I)
                    s = re.sub(r'\s+', ' ', s).strip()
                    return s

                # Build counts keyed by normalized title and by normalized title|company when company exists
                title_to_sector_counts: Dict[str, Dict[str, int]] = {}
                for r in rows:
                    if not isinstance(r, (list, tuple)):
                        continue
                    tval = ""
                    sval = ""
                    cval = ""
                    if title_idx is not None and title_idx < len(r):
                        tval = r[title_idx] if isinstance(r[title_idx], str) else (str(r[title_idx]) if r[title_idx] is not None else "")
                        tval = tval.strip()
                    if sector_idx is not None and sector_idx < len(r):
                        sval = r[sector_idx] if isinstance(r[sector_idx], str) else (str(r[sector_idx]) if r[sector_idx] is not None else "")
                        sval = sval.strip()
                    if company_idx is not None and company_idx < len(r):
                        cval = r[company_idx] if isinstance(r[company_idx], str) else (str(r[company_idx]) if r[company_idx] is not None else "")
                        cval = cval.strip()
                    if not tval or not sval:
                        continue
                    n_title = _normalize_title_for_matching(tval)
                    n_company = (cval or "").strip().lower()
                    # company-specific key
                    if n_company:
                        key_cc = f"{n_title}|{n_company}"
                        d = title_to_sector_counts.setdefault(key_cc, {})
                        d[sval] = d.get(sval, 0) + 1
                    # title-only key
                    key_t = n_title
                    d2 = title_to_sector_counts.setdefault(key_t, {})
                    d2[sval] = d2.get(sval, 0) + 1

                # Apply to per_map entries: try title+company first, then title
                applied = 0
                for entry in per_map:
                    try:
                        entry_title = (entry.get("title") or "").strip()
                        if not entry_title:
                            continue
                        entry_company = (entry.get("company") or "").strip().lower()
                        n_entry_title = _normalize_title_for_matching(entry_title)
                        selected_sector = ""
                        if entry_company:
                            kcc = f"{n_entry_title}|{entry_company}"
                            if kcc in title_to_sector_counts:
                                sorted_s = sorted(title_to_sector_counts[kcc].items(), key=lambda x: x[1], reverse=True)
                                if sorted_s:
                                    selected_sector = sorted_s[0][0]
                        if not selected_sector and n_entry_title in title_to_sector_counts:
                            sorted_s2 = sorted(title_to_sector_counts[n_entry_title].items(), key=lambda x: x[1], reverse=True)
                            if sorted_s2:
                                selected_sector = sorted_s2[0][0]
                        existing_sector = (entry.get("sector") or "").strip()
                        if selected_sector and not existing_sector:
                            entry["sector"] = selected_sector
                            applied += 1
                    except Exception:
                        # safe continue for per-entry issues
                        continue

                # Debug print to assist diagnosis (non-fatal)
                try:
                    total_keys = len(title_to_sector_counts)
                    total_rows = len(rows)
                    print(f"[DataSorter] sector-mapping: rows={total_rows} keys={total_keys} applied={applied}")
                except Exception:
                    pass
    except Exception:
        import traceback
        traceback.print_exc()
        # non-fatal; continue

    # NEW: Build title -> company mapping and enrich per_map with company & sector info,
    # Keep Seniority and Company separate (do NOT append company into seniority or inject sector into family).
    try:
        if fetch_from_db and uploaded_columns:
            lower_headers = [str(h).lower() for h in uploaded_columns]
            if 'company' in lower_headers:
                title_to_company_counts: Dict[str, Dict[str, int]] = {}
                title_idx = detect_title_column(uploaded_columns)
                company_idx = None
                for i,h in enumerate(uploaded_columns):
                    try:
                        if isinstance(h, str) and h.strip().lower() == 'company':
                            company_idx = i
                            break
                    except Exception:
                        continue
                if title_idx is not None and company_idx is not None:
                    for r in rows:
                        if not isinstance(r, (list, tuple)):
                            continue
                        tval = ""
                        cval = ""
                        if title_idx is not None and title_idx < len(r):
                            tval = r[title_idx] if isinstance(r[title_idx], str) else (str(r[title_idx]) if r[title_idx] is not None else "")
                            tval = tval.strip()
                        if company_idx is not None and company_idx < len(r):
                            cval = r[company_idx] if isinstance(r[company_idx], str) else (str(r[company_idx]) if r[company_idx] is not None else "")
                            cval = cval.strip()
                        if not tval or not cval:
                            continue
                        tl = tval.lower()
                        title_to_company_counts.setdefault(tl, {})
                        title_to_company_counts[tl][cval] = title_to_company_counts[tl].get(cval, 0) + 1

                    # sector hint from payload or project context
                    sector_hint = (data.get('sector') or "").strip() or (project_context.get('sector') if project_context else "") or ""

                    for title_lower, counts in title_to_company_counts.items():
                        entry = title_entry_map.get(title_lower)
                        if not entry:
                            continue
                        sorted_c = sorted(counts.items(), key=lambda x: x[1], reverse=True)
                        company_choice = sorted_c[0][0] if sorted_c else ""
                        # attach most common company for this title (as a separate field)
                        if company_choice:
                            entry["company"] = company_choice
                        # ensure sector present on entry, prefer existing sector -> db sector -> inferred by company/family
                        existing_sector = (entry.get("sector") or "").strip()
                        if not existing_sector:
                            inferred_sector = ""
                            try:
                                inferred_sector = infer_company_sector(company_choice or "", entry.get("family") or "", entry.get("country") or "", model_name=gemini_model) or ""
                            except Exception:
                                inferred_sector = ""
                            entry["sector"] = inferred_sector or sector_hint or ""
                        # Ensure jobfamily key exists separately if family is present
                        if ("jobfamily" not in entry or not entry.get("jobfamily")) and entry.get("family"):
                            entry["jobfamily"] = entry.get("family")
                        # IMPORTANT: Do NOT combine company into seniority or append sector into family string.
    except Exception:
        import traceback
        traceback.print_exc()
        # Non-fatal: proceed without company enrichment if anything fails

    # --- AFFECTED section: add skillset + sector persistence after base enrichment updates ---
    if fetch_from_db:
        try:
            import psycopg2
            from psycopg2 import sql
            pg_host=os.getenv("PGHOST","localhost")
            pg_port=int(os.getenv("PGPORT","5432"))
            pg_user=os.getenv("PGUSER","postgres")
            pg_password=os.getenv("PGPASSWORD","") or "orlha"
            pg_db=os.getenv("PGDATABASE","candidate_db")
            conn_up = psycopg2.connect(host=pg_host, port=pg_port, user=pg_user, password=pg_password, dbname=pg_db)
            cur_up = conn_up.cursor()

            # Update jobfamily/seniority/geographic first (unchanged original logic)
            col_indices = {h: idx for idx, h in enumerate(uploaded_columns)}
            title_col_name = None
            for h in uploaded_columns:
                if isinstance(h, str) and h.lower() in ('jobtitle','role'):
                    title_col_name = h
                    break
            if not title_col_name:
                for h in uploaded_columns:
                    if isinstance(h, str) and ('title' in h.lower() or 'role' in h.lower()):
                        title_col_name = h
                        break
            for row_idx, norm_row in enumerate(rows):
                try:
                    id_val = None
                    userid_val = None
                    username_val = None
                    name_val = None
                    company_val = None
                    title_val = ""
                    ctid_val = None
                    if isinstance(norm_row, (list, tuple)):
                        if 'id' in col_indices and col_indices['id'] < len(norm_row):
                            id_val = norm_row[col_indices['id']]
                        if 'userid' in col_indices and col_indices['userid'] < len(norm_row):
                            userid_val = norm_row[col_indices['userid']]
                        if 'username' in col_indices and col_indices['username'] < len(norm_row):
                            username_val = norm_row[col_indices['username']]
                        if 'name' in col_indices and col_indices['name'] < len(norm_row):
                            name_val = norm_row[col_indices['name']]
                        if 'company' in col_indices and col_indices['company'] < len(norm_row):
                            company_val = norm_row[col_indices['company']]
                        if '_ctid' in col_indices and col_indices['_ctid'] < len(norm_row):
                            ctid_val = norm_row[col_indices['_ctid']]
                        t_idx = detect_title_column(uploaded_columns)
                        if t_idx is not None and t_idx < len(norm_row):
                            title_val = norm_row[t_idx] if isinstance(norm_row[t_idx], str) else (str(norm_row[t_idx]) if norm_row[t_idx] is not None else "")
                            title_val = title_val.strip()
                    else:
                        continue
                    if not title_val:
                        continue
                    entry = title_entry_map.get(title_val.lower())
                    if not entry:
                        continue
                    jf_val = entry.get('jobfamily') or entry.get('family') or ""
                    sen_val = entry.get('seniority') or ""
                    geo_val = entry.get('geographic') or entry.get('region') or ""
                    set_parts = []
                    params = []
                    if jf_val:
                        set_parts.append(sql.SQL("jobfamily = %s"))
                        params.append(jf_val)
                    if sen_val:
                        set_parts.append(sql.SQL("seniority = %s"))
                        params.append(sen_val)
                    if geo_val:
                        set_parts.append(sql.SQL("geographic = %s"))
                        params.append(geo_val)
                    if not set_parts:
                        continue
                    where_parts = []
                    where_params = []
                    if id_val is not None and str(id_val) != "":
                        where_parts = [sql.SQL("id = %s")]
                        where_params = [id_val]
                    elif ctid_val:
                        where_parts = [sql.SQL("ctid::text = %s")]
                        where_params = [str(ctid_val)]
                    else:
                        if userid_val:
                            where_parts.append(sql.SQL("userid = %s")); where_params.append(str(userid_val))
                        if username_val:
                            where_parts.append(sql.SQL("username = %s")); where_params.append(username_val)
                        if name_val:
                            where_parts.append(sql.SQL("name = %s")); where_params.append(name_val)
                        if company_val:
                            where_parts.append(sql.SQL("company = %s")); where_params.append(company_val)
                        if title_col_name:
                            where_parts.append(sql.SQL("{} = %s").format(sql.Identifier(title_col_name))); where_params.append(title_val)
                    if not where_parts:
                        continue
                    where_sql = sql.SQL(" AND ").join(where_parts)
                    query = sql.SQL("UPDATE process SET ") + sql.SQL(", ").join(set_parts) + sql.SQL(" WHERE ") + where_sql
                    cur_up.execute(query.as_string(conn_up), params + where_params)
                except Exception:
                    import traceback
                    traceback.print_exc()
                    continue
            conn_up.commit()

            # Skillset + sector persistence (new addition)
            try:
                org_header = find_existing_org_header(uploaded_columns) or ("company" if "company" in uploaded_columns else None)
                per_by_title = { (p.get("title") or "").strip().lower(): p for p in per_map }
                for row_idx, norm_row in enumerate(rows):
                    try:
                        t_idx = detect_title_column(uploaded_columns)
                        if t_idx is None or t_idx >= len(norm_row):
                            continue
                        jobtitle = str(norm_row[t_idx] or "").strip()
                        if not jobtitle:
                            continue
                        title_lower = jobtitle.lower()
                        company_val = ""
                        if org_header and org_header in uploaded_columns:
                            ci = uploaded_columns.index(org_header)
                            if ci < len(norm_row):
                                company_val = str(norm_row[ci] or "").strip()
                        country_val = ""
                        if "country" in uploaded_columns:
                            ci = uploaded_columns.index("country")
                            if ci < len(norm_row):
                                country_val = str(norm_row[ci] or "").strip()
                        entry = per_by_title.get(title_lower) or {}
                        jf_val = entry.get("jobfamily") or entry.get("family") or ""
                        sen_val = entry.get("seniority") or ""
                        cache_key = (
                            jobtitle.strip().lower(),
                            (company_val or "").strip().lower(),
                            "",
                            (country_val or "").strip().lower(),
                            (jf_val or "").strip().lower(),
                            (sen_val or "").strip().lower()
                        )
                        skills = _SKILLSET_CACHE.get(cache_key)
                        if skills is None:
                            skills = fetch_gemini_skillset(
                                jobtitle=jobtitle,
                                company=company_val,
                                project_title="",
                                country=country_val,
                                job_family=jf_val,
                                seniority=sen_val,
                                model_name=gemini_model
                            )
                            _SKILLSET_CACHE[cache_key] = skills
                        skillset_str = "; ".join(skills) if skills else ""
                        sector_inferred = ""
                        # Use existing sector cell if populated
                        if "sector" in uploaded_columns:
                            si = uploaded_columns.index("sector")
                            if si < len(norm_row):
                                existing_sector_cell = str(norm_row[si] or "").strip()
                                if existing_sector_cell:
                                    sector_inferred = existing_sector_cell
                        if not sector_inferred:
                            sector_inferred = infer_company_sector(company_val, jf_val, country_val, model_name=gemini_model)
                        if not sector_inferred:
                            sector_inferred = sector

                        set_parts = []
                        params = []
                        if skillset_str:
                            set_parts.append(sql.SQL("skillset = %s"))
                            params.append(skillset_str)
                        if sector_inferred:
                            set_parts.append(sql.SQL("sector = %s"))
                            params.append(sector_inferred)
                        if not set_parts:
                            continue

                        where_parts = []
                        where_params = []
                        id_val = None
                        ctid_val = None
                        if "id" in uploaded_columns:
                            ci = uploaded_columns.index("id")
                            if ci < len(norm_row):
                                id_val = norm_row[ci]
                        if "_ctid" in uploaded_columns:
                            ci = uploaded_columns.index("_ctid")
                            if ci < len(norm_row):
                                ctid_val = norm_row[ci]

                        if id_val is not None and str(id_val) != "":
                            where_parts = [sql.SQL("id = %s")]
                            where_params = [id_val]
                        elif ctid_val:
                            where_parts = [sql.SQL("ctid::text = %s")]
                            where_params = [str(ctid_val)]
                        else:
                            if "userid" in uploaded_columns:
                                ci = uploaded_columns.index("userid")
                                if ci < len(norm_row) and norm_row[ci]:
                                    where_parts.append(sql.SQL("userid = %s")); where_params.append(str(norm_row[ci]))
                            if "username" in uploaded_columns:
                                ci = uploaded_columns.index("username")
                                if ci < len(norm_row) and norm_row[ci]:
                                    where_parts.append(sql.SQL("username = %s")); where_params.append(norm_row[ci])
                            if "name" in uploaded_columns:
                                ci = uploaded_columns.index("name")
                                if ci < len(norm_row) and norm_row[ci]:
                                    where_parts.append(sql.SQL("name = %s")); where_params.append(norm_row[ci])
                            if "company" in uploaded_columns:
                                ci = uploaded_columns.index("company")
                                if ci < len(norm_row) and norm_row[ci]:
                                    where_parts.append(sql.SQL("company = %s")); where_params.append(norm_row[ci])
                            title_col_name = None
                            for h in uploaded_columns:
                                if isinstance(h, str) and (h.lower() in ('jobtitle','role') or 'title' in h.lower() or 'role' in h.lower()):
                                    title_col_name = h
                                    break
                            if title_col_name:
                                where_parts.append(sql.SQL("{} = %s").format(sql.Identifier(title_col_name))); where_params.append(jobtitle)
                        if not where_parts:
                            continue

                        query = sql.SQL("UPDATE process SET ") + sql.SQL(", ").join(set_parts) + sql.SQL(" WHERE ") + sql.SQL(" AND ").join(where_parts)
                        cur_up.execute(query.as_string(conn_up), params + where_params)
                    except Exception:
                        import traceback
                        traceback.print_exc()
                        continue
                conn_up.commit()
            except Exception:
                import traceback
                traceback.print_exc()

            cur_up.close()
            conn_up.close()
        except Exception:
            import traceback
            traceback.print_exc()
    # --- END AFFECTED section ---

    # Merge job family with sector on per_map so frontend sees combined information.
    # We perform this merge here (after sector and family have been set above) so it's present
    # in the JSON returned to the client and in subsequent export logic.
    try:
        merged_count = 0
        for rec in per_map:
            try:
                fam = (rec.get("family") or rec.get("jobfamily") or "").strip()
                sec = (rec.get("sector") or "").strip()
                # if both exist, merge into jobfamily as "Sector - Family" unless already present
                if sec and fam:
                    merged = f"{sec} - {fam}"
                    # avoid double-merging if fam already contains sector
                    if sec.lower() not in fam.lower():
                        rec["jobfamily"] = merged
                        merged_count += 1
                    else:
                        # ensure jobfamily key exists
                        if not rec.get("jobfamily"):
                            rec["jobfamily"] = fam
                else:
                    # ensure jobfamily exists when family exists
                    if fam and not rec.get("jobfamily"):
                        rec["jobfamily"] = fam
            except Exception:
                continue
        # small debug print to help troubleshooting without changing behavior
        try:
            print(f"[DataSorter] Merged sector+family for {merged_count} per_map entries")
        except Exception:
            pass
    except Exception:
        import traceback
        traceback.print_exc()

    allowed_sector_fams = _resolve_allowed_families(sector_norm)
    if allowed_sector_fams:
        allowed_set = set(allowed_sector_fams)
        for rec in per_map:
            if rec.get("family") and rec["family"] not in allowed_set:
                rec["family"] = None
    location_annotations=[]
    for row_index,ov in origin_overrides.items():
        if row_index < 0 or row_index >= len(titles):
            continue
        location_annotations.append({
            "row_index": row_index,
            "title": titles[row_index],
            "annotation": ov["annotation"]
        })
    for row_index,ov in role_overrides.items():
        if row_index < 0 or row_index >= len(titles):
            continue
        location_annotations.append({
            "row_index": row_index,
            "title": titles[row_index],
            "annotation": ov["annotation"]
        })
    regions=[p.get("geographic","") for p in per_map]
    countries=[p.get("country","") for p in per_map]
    geo_counts={}
    for g in regions:
        if g:
            geo_counts[g]=geo_counts.get(g,0)+1
    cty_counts={}
    for c in countries:
        if c:
            cty_counts[c]=cty_counts.get(c,0)+1
    geo_total=sum(geo_counts.values()) or 1
    cty_total=sum(cty_counts.values()) or 1
    geographic_suggestions=[{
        "geographic": g,
        "count": cnt,
        "coverage": round(cnt/geo_total,4)
    } for g,cnt in sorted(geo_counts.items(), key=lambda x: x[1], reverse=True)]
    country_suggestions=[{
        "country": c,
        "count": cnt,
        "coverage": round(cnt/cty_total,4)
    } for c,cnt in sorted(cty_counts.items(), key=lambda x: x[1], reverse=True)]
    elapsed=time.time()-start
    print(f"[DataSorter] Inference {len(titles)} titles in {elapsed:.2f}s | sector={sector_norm or 'N/A'} origin_rows={len(origin_overrides)} role_rows={len(role_overrides)}")
    response={
        **result,
        "processing_seconds": elapsed,
        "title_count": len([t for t in titles if t]),
        "geographic_suggestions": geographic_suggestions,
        "country_suggestions": country_suggestions,
        "project_context": project_context,
        "default_region": default_region or "",
        "default_country": default_country or "",
        "origin_overrides": [{"row_index": ri, **ov} for ri,ov in origin_overrides.items()],
        "role_overrides": [{"row_index": ri, **ov} for ri,ov in role_overrides.items()],
        "location_annotations": location_annotations,
        "rendering_count": rendering_count
    }
    return jsonify(response)

@app.route('/update_process_enrichment', methods=['POST'])
def update_process_enrichment():
    data = request.get_json(force=True) or {}
    entries = data.get('entries') or data.get('rows') or []
    if not isinstance(entries, list) or not entries:
        return jsonify({"ok": False, "error": "No entries provided"}), 400

    try:
        import psycopg2
        from psycopg2 import sql
        pg_host=os.getenv("PGHOST","localhost")
        pg_port=int(os.getenv("PGPORT","5432"))
        pg_user=os.getenv("PGUSER","postgres")
        pg_password=os.getenv("PGPASSWORD","") or "orlha"
        pg_db=os.getenv("PGDATABASE","candidate_db")
        conn=psycopg2.connect(host=pg_host, port=pg_port, user=pg_user, password=pg_password, dbname=pg_db)
        cur=conn.cursor()

        cur.execute("""
            SELECT column_name
            FROM information_schema.columns
            WHERE table_schema='public' AND table_name='process'
        """)
        proc_cols = [r[0].lower() for r in cur.fetchall()]
        preferred_title_col = None
        if 'jobtitle' in proc_cols:
            preferred_title_col = 'jobtitle'
        elif 'role' in proc_cols:
            preferred_title_col = 'role'
        else:
            for c in proc_cols:
                if 'title' in c or 'role' in c:
                    preferred_title_col = c
                    break

        total_updated = 0
        errors = []
        for idx, entry in enumerate(entries):
            try:
                jf = entry.get('jobfamily') if entry.get('jobfamily') is not None else entry.get('family') or None
                geo = entry.get('geographic') if entry.get('geographic') is not None else entry.get('region') or None
                sen = entry.get('seniority') if entry.get('seniority') is not None else None

                set_parts = []
                params = []
                if jf is not None:
                    set_parts.append(sql.SQL("jobfamily = %s"))
                    params.append(jf)
                if geo is not None:
                    set_parts.append(sql.SQL("geographic = %s"))
                    params.append(geo)
                if sen is not None:
                    set_parts.append(sql.SQL("seniority = %s"))
                    params.append(sen)

                if not set_parts:
                    continue

                where_sql = None
                where_params = []
                if entry.get('userid') is not None:
                    where_sql = sql.SQL("userid = %s")
                    where_params = [str(entry.get('userid'))]
                elif entry.get('username') is not None:
                    where_sql = sql.SQL("username = %s")
                    where_params = [entry.get('username')]
                else:
                    name = entry.get('name') or entry.get('candidate') or ""
                    company = entry.get('company') or ""
                    title_val = entry.get('jobtitle') or entry.get('title') or ""
                    if not preferred_title_col:
                        raise ValueError("No title column available in process table to match entry without userid/username")
                    where_parts = []
                    where_params = []
                    if name:
                        where_parts.append(sql.SQL("name = %s"))
                        where_params.append(name)
                    if company:
                        where_parts.append(sql.SQL("company = %s"))
                        where_params.append(company)
                    if title_val:
                        where_parts.append(sql.SQL("{} = %s").format(sql.Identifier(preferred_title_col)))
                        where_params.append(title_val)
                    if not where_parts:
                        raise ValueError("Insufficient identifiers to locate process row (provide userid, username, or name/company/jobtitle)")
                    where_sql = sql.SQL(" AND ").join(where_parts)

                query = sql.SQL("UPDATE process SET ") + sql.SQL(", ").join(set_parts) + sql.SQL(" WHERE ") + where_sql
                exec_params = params + where_params
                cur.execute(query.as_string(conn), exec_params)
                total_updated += cur.rowcount if cur.rowcount is not None else 0
            except Exception as e:
                errors.append({"index": idx, "error": str(e), "entry": entry})
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({"ok": True, "updated": total_updated, "errors": errors})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route('/infer_project_date', methods=['POST'])
def infer_project_date_route():
    data = request.get_json(force=True)
    company = (data.get("company") or "").strip()
    product = (data.get("product") or "").strip()
    sector = (data.get("sector") or "").strip().lower()
    if sector != "gaming" or not company or not product:
        return jsonify({"project_date": ""})
    date_val = fetch_gemini_project_date(company, product)
    return jsonify({"project_date": date_val})

@app.route('/generate_excel', methods=['POST'])
def generate_excel_route():
    data=request.get_json(force=True)
    try:
        # ===== Affected section: update JobFamilyRoles in static/data_sorter.json using uploaded titles =====
        uploaded_cols = data.get('uploaded_columns', []) or []
        rows = data.get('rows', []) or []
        # Normalize rows if dicts; reuse existing helper
        rows = normalize_rows(uploaded_cols, rows)
        titles = extract_titles(uploaded_cols, rows)

        def _normalize_core_title(raw: str) -> str:
            if not raw or not isinstance(raw, str):
                return ""
            t = raw.strip()
            # strip parentheses content
            import re
            t = re.sub(r'\(.*?\)', ' ', t)
            # remove designation tokens
            t = re.sub(r'\b(?:senior|sr\.?|jr\.?|junior|lead|principal|staff|associate|assistant|manager|director|vp\b|vice\b|president|chief|head|executive|intern|trainee)\b', ' ', t, flags=re.I)
            # remove ordinals/roman numerals and numeric levels
            t = re.sub(r'\b(?:i{1,3}|[ivx]+|[0-9]+(?:st|nd|rd|th)?)\b', ' ', t, flags=re.I)
            # cleanup punctuation/stopwords and whitespace
            t = re.sub(r'[^\w&/+\- ]', ' ', t)
            t = re.sub(r'\b(of|the|and|for)\b', ' ', t, flags=re.I)
            t = re.sub(r'\s+', ' ', t).strip()
            return t.title()

        def _heuristic_family(title: str) -> str:
            tl = (title or "").lower()
            if any(k in tl for k in ('engineer','developer','software','programmer','devops','backend','frontend')):
                return "Programming"
            if any(k in tl for k in ('artist','animation','animator','vfx','lighting','3d','concept art','technical artist')):
                return "Animation" if ('anim' in tl or 'animation' in tl) else "Art"
            if any(k in tl for k in ('designer','design','ux','ui','level','narrative','game designer')):
                return "Design"
            if any(k in tl for k in ('qa','tester','quality assurance','test engineer')):
                return "QA"
            if any(k in tl for k in ('audio','sound','composer')):
                return "Audio"
            if any(k in tl for k in ('producer','production','project manager','project coordinator')):
                return "Production"
            if any(k in tl for k in ('data scientist','data analyst','ml','machine learning','ai','ai engineer')):
                return "Data Science"
            if any(k in tl for k in ('finance','financial','accountant','fp&a','controller')):
                return "Finance"
            return "Corporate"

        # Load current config
        cfg_path = os.path.join(app.static_folder or 'static', 'data_sorter.json')
        try:
            with open(cfg_path, encoding='utf-8') as f:
                cfg = json.load(f) or {}
        except Exception:
            cfg = {}
        jobfamilyroles = cfg.get("JobFamilyRoles") or {}

        # Build set of existing titles (case-insensitive)
        existing_lower = set()
        for fam, fam_titles in jobfamilyroles.items():
            if isinstance(fam_titles, list):
                for t in fam_titles:
                    if isinstance(t, str) and t.strip():
                        existing_lower.add(t.strip().lower())

        # Prepare additions by core title uniqueness
        cores = []
        seen_core = set()
        for t in titles or []:
            core = _normalize_core_title(t or "")
            if core and core.strip().lower() not in seen_core:
                seen_core.add(core.strip().lower())
                cores.append(core)

        additions = []
        for core in cores:
            if core.strip().lower() in existing_lower:
                continue
            fam = _heuristic_family(core)
            jobfamilyroles.setdefault(fam, [])
            # avoid duplicate within family
            fam_existing_lower = { (x or "").strip().lower() for x in jobfamilyroles.get(fam, []) if isinstance(x, str) }
            if core.strip().lower() not in fam_existing_lower:
                jobfamilyroles[fam].append(core)
                additions.append((fam, core))

        if additions:
            cfg["JobFamilyRoles"] = jobfamilyroles
            # Safe overwrite: write temp then replace
            tmp_path = cfg_path + ".tmp"
            try:
                with open(tmp_path, "w", encoding="utf-8") as f:
                    json.dump(cfg, f, indent=2, ensure_ascii=False)
                os.replace(tmp_path, cfg_path)
                # Log the exact titles added
                print(f"[JobFamilies] Updated data_sorter.json with {len(additions)} new titles: {', '.join([title for _, title in additions])}")
            except Exception as e:
                try:
                    if os.path.exists(tmp_path):
                        os.remove(tmp_path)
                except Exception:
                    pass
                print(f"[JobFamilies] Failed to update data_sorter.json: {e}")

            # Append RecentUpdates entries so frontend can identify new ones clearly
            try:
                recent = cfg.get("RecentUpdates", [])
                import time as _t
                for fam, title in additions:
                    recent_entry = {
                        "ts": int(_t.time()),
                        "iso": _t.strftime("%Y-%m-%dT%H:%M:%SZ", _t.gmtime()),
                        "action": "Added",
                        "family": fam,
                        "title": title
                    }
                    recent.insert(0, recent_entry)
                cfg["RecentUpdates"] = recent[:100]
                # Write JSON again to include RecentUpdates
                tmp_path2 = cfg_path + ".tmp"
                with open(tmp_path2, "w", encoding="utf-8") as f2:
                    json.dump(cfg, f2, indent=2, ensure_ascii=False)
                os.replace(tmp_path2, cfg_path)
                print(f"[JobFamilies] RecentUpdates appended for {len(additions)} new titles: {', '.join([t for _, t in additions])}")
            except Exception as e:
                print(f"[JobFamilies] Failed to append RecentUpdates: {e}")
        # ===== End affected section: JobFamilyRoles overwrite =====

        # Generate Excel strictly from process table for the current user with required columns only (including sourcingstatus restoration)
        try:
            import psycopg2
            from psycopg2 import sql
            pg_host=os.getenv("PGHOST","localhost")
            pg_port=int(os.getenv("PGPORT","5432"))
            pg_user=os.getenv("PGUSER","postgres")
            pg_password=os.getenv("PGPASSWORD","") or "orlha"
            pg_db=os.getenv("PGDATABASE","candidate_db")
            conn = psycopg2.connect(host=pg_host, port=pg_port, user=pg_user, password=pg_password, dbname=pg_db)
            cur = conn.cursor()
            # Base required columns
            target_cols = ["name","company","jobtitle","country","linkedinurl","sector","jobfamily","geographic","seniority","skillset"]
            # Detect if sourcingstatus exists in process table; if so include it (restoring previous presence)
            cur.execute("SELECT column_name FROM information_schema.columns WHERE table_schema='public' AND table_name='process'")
            proc_cols = {r[0].lower() for r in cur.fetchall()}
            if 'sourcingstatus' in proc_cols and 'sourcingstatus' not in target_cols:
                target_cols.append('sourcingstatus')

            # Build WHERE from session or payload
            where_clauses = []
            params = []
            payload_userid = data.get('userid')
            payload_username = data.get('username')
            sess_userid = session.get('userid')
            sess_username = session.get('username')
            userid = payload_userid if payload_userid is not None else sess_userid
            username = payload_username if payload_username is not None else sess_username
            if userid:
                where_clauses.append(sql.SQL("userid = %s")); params.append(str(userid))
            elif username:
                where_clauses.append(sql.SQL("username = %s")); params.append(username)
            else:
                cur.close(); conn.close()
                return jsonify({"error":"Missing userid/username to filter process table"}), 400

            select_fields = sql.SQL(', ').join([sql.Identifier(c) for c in target_cols])
            base_query = sql.SQL("SELECT {fields} FROM process").format(fields=select_fields)
            where_sql = sql.SQL(" WHERE ") + sql.SQL(" AND ").join(where_clauses)
            query = base_query + where_sql + sql.SQL(" ORDER BY name NULLS LAST")
            cur.execute(query.as_string(conn), params)
            fetched = cur.fetchall()
            cur.close(); conn.close()

            # Create Excel with required columns (including restored sourcingstatus if available)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Selections"
            ws.append(target_cols)
            for r in fetched:
                row_dict = dict(zip(target_cols, r))
                ws.append([row_dict.get(c, "") for c in target_cols])

            # Restore sourcingstatus dropdown on the generated Excel if the column is present
            if "sourcingstatus" in target_cols:
                sourcing_col_idx = target_cols.index("sourcingstatus") + 1  # 1-based
                col_letter = openpyxl.utils.get_column_letter(sourcing_col_idx)
                cell_range = f"{col_letter}2:{col_letter}{ws.max_row}"
                dropdown_values = data.get('sourcingstatus', [])
                joined = ",".join([str(v).replace(",", "").replace('"', "").replace("\n", "").strip() for v in dropdown_values if v])
                if joined:
                    if len(joined) <= 255:
                        dv = DataValidation(type="list", formula1=f'"{joined}"', allow_blank=True)
                        ws.add_data_validation(dv)
                        dv.add(cell_range)
                    else:
                        list_sheet = "Lists"
                        if list_sheet not in wb.sheetnames:
                            ls = wb.create_sheet(list_sheet)
                        else:
                            ls = wb[list_sheet]
                        write_col = 1
                        while ls.cell(row=1, column=write_col).value is not None:
                            write_col += 1
                        cleaned_vals = []
                        seen = set()
                        for v in dropdown_values:
                            if not v:
                                continue
                            c = str(v).replace(",", "").replace('"', "").replace("\n", "").strip()
                            if c and c not in seen:
                                seen.add(c)
                                cleaned_vals.append(c)
                        for i, val in enumerate(cleaned_vals, 1):
                            ls.cell(row=i, column=write_col, value=val)
                        letter = openpyxl.utils.get_column_letter(write_col)
                        ref = f"{list_sheet}!${letter}$1:${letter}${len(cleaned_vals)}"
                        dv = DataValidation(type="list", formula1=f"={ref}", allow_blank=True)
                        ws.add_data_validation(dv)
                        dv.add(cell_range)

            tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            wb.save(tmp.name)
            tmp.close()
            file_path = tmp.name
        except Exception as e:
            import traceback
            traceback.print_exc()
            return jsonify({"error": f"Process table fetch for Excel failed: {e}"}), 500

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    response=send_file(file_path, as_attachment=True, download_name="user_selections.xlsx")
    @response.call_on_close
    def cleanup():
        try: os.remove(file_path)
        except OSError: pass
    # Log the human-triggered export action
    try:
        _log_approval_ds(
            action="export_excel_triggered",
            username=session.get('username', ''),
            detail="generate_excel export completed",
            source="data_sorter.py",
        )
    except Exception:
        pass
    return response

def _serve_html_fallback(filename: str):
    candidates = [
        app.static_folder,
        os.path.join(app.root_path, 'static'),
        os.getcwd(),
        app.root_path
    ]
    for folder in candidates:
        if not folder:
            continue
        path = os.path.join(folder, filename)
        if os.path.exists(path):
            print(f"[serve] Serving {filename} from {folder}")
            return send_from_directory(folder, filename)
    msg = (f"<h2>404 - {filename} not found</h2>"
           f"<p>Expected {filename} in one of these locations: {', '.join(candidates)}.</p>")
    return msg, 404

@app.route('/login.html', methods=['GET'])
def serve_login_html():
    return _serve_html_fallback('login.html')

@app.route('/login', methods=['GET'])
def serve_login_get():
    return _serve_html_fallback('login.html')

@app.route('/register.html', methods=['GET'])
def serve_register_html():
    return _serve_html_fallback('register.html')

@app.route('/user/resolve', methods=['GET'])
def user_resolve_route():
    try:
        uid = session.get('userid')
        uname = session.get('username')
        full = session.get('full_name', '') if session.get('full_name') else ""
        if uid and uname:
            return jsonify({"ok": True, "userid": uid, "username": uname, "full_name": full})
    except Exception:
        pass
    qname = request.args.get('username')
    if qname:
        user = _fetch_user_by_username(qname)
        if user:
            return jsonify({
                "ok": True,
                "userid": user.get('id') or user.get('username'),
                "username": user.get('username'),
                "full_name": user.get('full_name') or ""
            })
        return jsonify({"ok": False}), 404
    return jsonify({"ok": False}), 404

# Serve data_sorter.html explicitly (use fallback helper if file missing)
@app.route('/data_sorter.html', methods=['GET'])
def serve_data_sorter_html():
    return _serve_html_fallback('data_sorter.html')

@app.route('/')
def index():
    # Serve the main UI HTML (data_sorter.html). When embedded via ?embed=1,
    # X-Embed header, or when the browser sets Sec-Fetch-Dest: iframe, authentication
    # will be bypassed as implemented in require_login().
    # Use the fallback helper which will look in static and nearby locations.
    return _serve_html_fallback('data_sorter.html')

if __name__ == '__main__':
    port = int(os.getenv("PORT", "8091"))
    app.run(debug=True, port=port)

@app.errorhandler(500)
def _handle_500_ds(e):
    _log_error_ds(source="data_sorter", message=str(e), severity="critical",
                  endpoint=request.path if request else "")
    return jsonify({"error": "Internal server error"}), 500

@app.after_request
def _capture_http_errors_ds(response):
    """Log any HTTP 4xx/5xx response to the Error Capture log."""
    _skip = ("/favicon.ico",)
    if (response.status_code >= 400
            and request.method != "OPTIONS"
            and not any(request.path.startswith(p) for p in _skip)):
        _sc = response.status_code
        _sev = "critical" if _sc >= 500 else "warning"
        _body_msg = ""
        try:
            if "json" in (response.content_type or ""):
                _body_msg = response.get_data(as_text=True)[:500]
        except Exception:
            pass
        _log_error_ds(
            source="data_sorter",
            message=f"{request.method} {request.path} → HTTP {_sc}",
            severity=_sev,
            endpoint=request.path,
            http_status=_sc,
            detail=_body_msg,
        )
    return response