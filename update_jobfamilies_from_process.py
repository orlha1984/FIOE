#!/usr/bin/env python3
"""
update_jobfamilies_from_process.py

Standalone utility to enrich / update the "JobFamilyRoles" section in static/data_sorter.json
using job titles found in a process table (database) or a local file (CSV, XLSX, JSON).

Behavior:
- Loads static/data_sorter.json (overwrites in place when --commit used).
- Reads job titles from either:
    * Postgres process table (default, when --db is provided), or
    * A local file (--file / -f) in CSV, XLSX, or JSON array-of-objects format.
- Normalizes titles by removing seniority/designation tokens (Lead, Senior, Junior, Manager, Director, etc.)
  before comparing/adding.
- Uses title2vec_service.map_titles_to_families_and_seniority() to propose a family for new core titles.
- Applies simple heuristics fallback if the model can't pick a family.
- Avoids duplicates (case-insensitive).
- Overwrites static/data_sorter.json only when --commit is specified (otherwise prints a dry-run summary).

Usage:
  python update_jobfamilies_from_process.py --file my_process_table.csv --commit
  python update_jobfamilies_from_process.py --db --userid 42 --limit 500 --commit
  python update_jobfamilies_from_process.py --file table.xlsx         (dry-run)

Notes:
- The script intentionally mirrors the DB env var conventions used by data_sorter.py:
    PGHOST, PGPORT, PGUSER, PGPASSWORD, PGDATABASE
- It requires the project layout where static/data_sorter.json is next to this script or in sibling 'static/'.
"""

import argparse
import csv
import json
import logging
import os
import re
import sys
import time
from typing import List, Dict, Set, Tuple, Optional

# logger for consistent server messages
logger = logging.getLogger("JobFamilies")
if not logger.handlers:
    # avoid reconfiguring global logging unnecessarily; ensure basic handler for standalone runs
    logging.basicConfig(level=logging.INFO)

# Optional heavy imports only when needed
try:
    from title2vec_service import map_titles_to_families_and_seniority  # reuses your existing classifier
except Exception:
    map_titles_to_families_and_seniority = None

# Excel support (already used in your project)
try:
    import openpyxl
except Exception:
    openpyxl = None

# DB support
try:
    import psycopg2
    from psycopg2 import sql
except Exception:
    psycopg2 = None
    sql = None

# Try to use gemini_context's helper if available for validation
try:
    import gemini_context  # module in the project that wraps Gemini calls
    _HAS_GEMINI = hasattr(gemini_context, "_gemini_generate_text")
except Exception:
    gemini_context = None
    _HAS_GEMINI = False

# Simple cache to avoid repeated LLM calls for same phrase in one run
_TITLE_VALIDATION_CACHE: Dict[str, bool] = {}

def is_likely_job_title_gemini(title: str, model_name: Optional[str] = None) -> bool:
    """
    Use Gemini (via gemini_context._gemini_generate_text) to ask a strict Yes/No:
    - "Is this phrase a job title as used in LinkedIn/resumes/HR systems? Answer 'Yes' or 'No' only."
    Returns True if model indicates 'Yes', False if 'No'.
    If Gemini is not available or an error occurs, returns True (safe permissive fallback).
    """
    if not title or not isinstance(title, str):
        return False
    key = title.strip().lower()
    if key in _TITLE_VALIDATION_CACHE:
        return _TITLE_VALIDATION_CACHE[key]

    # If Gemini helper not available, permissive fallback: allow addition (do not block)
    if not _HAS_GEMINI or gemini_context is None:
        _TITLE_VALIDATION_CACHE[key] = True
        return True

    # Build a concise, strict prompt asking for Yes/No only
    prompt = (
        'You are a strict assistant and must answer with only "Yes" or "No".\n'
        'Question: Is the following phrase a job title commonly used in professional resumes, LinkedIn, or HR systems?\n'
        f'Phrase: "{title}"\n'
        'If it is a department, discipline, team, project name, or otherwise not a job title, answer "No".'
    )
    try:
        # Use the project's wrapper which already configures the API key / model
        txt = gemini_context._gemini_generate_text(prompt, model_name=model_name, max_output_tokens=40)
        if not txt or not isinstance(txt, str):
            # treat absence as permissive (do not block) to avoid accidental loss
            result = True
        else:
            # Normalize and look for clear yes/no signals
            low = txt.strip().lower()
            yes = bool(re.search(r'\byes\b', low))
            no = bool(re.search(r'\bno\b', low))
            if yes and not no:
                result = True
            elif no and not yes:
                result = False
            else:
                # Heuristic: prefer initial token if it starts with yes/no
                if low.startswith('yes'):
                    result = True
                elif low.startswith('no'):
                    result = False
                else:
                    # Ambiguous -> conservative reject (treat as not a title)
                    # This is safer to avoid adding discipline-like phrases accidentally.
                    result = False
    except Exception:
        # On any exception contacting LLM, fall back permissively
        result = True

    _TITLE_VALIDATION_CACHE[key] = bool(result)
    return bool(result)

# -------------------- JobBERT local integration (option B) --------------------

# Per-run cache for JobBERT scores
_JOBBERT_SCORE_CACHE: Dict[str, Optional[float]] = {}
_JOBBERT_CLASSIFIER = None  # lazy-loaded transformers pipeline

# Thresholds and fallback mode (configurable via env)
JOBBERT_MODEL_ENV = os.getenv("VC_JOBBERT_MODEL", "TechWolf/JobBERT-v3")
JOBBERT_HIGH = float(os.getenv("JOBBERT_HIGH", "0.85"))
JOBBERT_LOW = float(os.getenv("JOBBERT_LOW", "0.20"))
JOBBERT_FALLBACK_MODE = os.getenv("JOBBERT_FALLBACK_MODE", "conservative").lower()  # 'conservative' or 'permissive'
_TRANSFORMERS_DEVICE = int(os.getenv("TRANSFORMERS_DEVICE", "-1"))  # -1 CPU, 0..G GPU

def _load_jobbert_classifier(model_id: str):
    """
    Lazy-load a local transformers text-classification pipeline for JobBERT.
    Returns pipeline or None on failure.
    """
    global _JOBBERT_CLASSIFIER
    if _JOBBERT_CLASSIFIER is not None:
        return _JOBBERT_CLASSIFIER if _JOBBERT_CLASSIFIER is not False else None
    try:
        # import here to avoid requirement if not used
        from transformers import pipeline
        # Use text-classification; many JobBERT variants are set up as classification models
        _JOBBERT_CLASSIFIER = pipeline("text-classification", model=model_id, device=_TRANSFORMERS_DEVICE)
        return _JOBBERT_CLASSIFIER
    except Exception:
        # Remember failure for this run
        _JOBBERT_CLASSIFIER = False
        return None

def jobbert_score_local(title: str, model_id: Optional[str] = None) -> Optional[float]:
    """
    Return a confidence score in [0.0, 1.0] that `title` is a job title using a local JobBERT model.
    Returns None if classifier unavailable or on error.
    """
    if not title or not isinstance(title, str):
        return None
    key = title.strip().lower()
    if key in _JOBBERT_SCORE_CACHE:
        return _JOBBERT_SCORE_CACHE[key]

    model_id = model_id or JOBBERT_MODEL_ENV
    clf = _load_jobbert_classifier(model_id)
    if not clf:
        _JOBBERT_SCORE_CACHE[key] = None
        return None

    try:
        # The pipeline may return a list of dicts with 'label' and 'score'
        out = clf(title, truncation=True)
        score = None
        if isinstance(out, list) and out:
            # Try to find a positive/title label
            for item in out:
                lbl = str(item.get("label", "")).lower()
                sc = float(item.get("score", 0.0) or 0.0)
                if any(k in lbl for k in ("title", "job", "yes", "positive")):
                    score = sc
                    break
            # Fallback: if only one result, use its score
            if score is None:
                try:
                    score = float(out[0].get("score", 0.0) or 0.0)
                except Exception:
                    score = None
        # Normalize to None or float
        if score is None:
            _JOBBERT_SCORE_CACHE[key] = None
            return None
        _JOBBERT_SCORE_CACHE[key] = float(score)
        return float(score)
    except Exception:
        # On classifier error, return None (unavailable)
        _JOBBERT_SCORE_CACHE[key] = None
        return None

# -------------------- End JobBERT integration --------------------

# Designation tokens to strip from titles when comparing core job title
_DESIGNATION_PATTERN = re.compile(
    r'\b(?:senior|sr\.?|jr\.?|junior|lead|principal|staff|associate|assistant|manager|director|vp\b|vice\b|president|chief|head|executive|intern|trainee)\b',
    flags=re.I
)
# Remove ordinal/version tokens (I, II, III, 1, 2) and parentheses content
_ORDINAL_PATTERN = re.compile(r'\b(?:i{1,3}|[ivx]+|[0-9]+(?:st|nd|rd|th)?)\b', flags=re.I)
_PAREN_PATTERN = re.compile(r'\(.*?\)')

# Title candidate token filters (for heuristic family fallback)
_PROGRAMMING_KEYWORDS = ('engineer', 'developer', 'software', 'programmer', 'devops', 'backend', 'frontend')
_ART_KEYWORDS = ('artist', 'animation', 'animator', 'vfx', 'lighting', '3d', 'concept art', 'technical artist')
_DESIGN_KEYWORDS = ('designer', 'design', 'ux', 'ui', 'level', 'narrative', 'game designer')
_QA_KEYWORDS = ('qa', 'tester', 'quality assurance', 'test engineer')
_AUDIO_KEYWORDS = ('audio', 'sound', 'composer')
_PROD_KEYWORDS = ('producer', 'production', 'project manager', 'project coordinator')
_DATA_KEYWORDS = ('data scientist', 'data analyst', 'ml', 'machine learning', 'ai', 'ai engineer')
_FINANCE_KEYWORDS = ('finance', 'financial', 'accountant', 'fp&a', 'controller')

# Path to the JSON file to update (relative to this script)
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_JSON_PATH = os.path.join(SCRIPT_DIR, "static", "data_sorter.json")


def find_title_column(headers: List[str]) -> Optional[int]:
    if not headers:
        return None
    kws = ["title", "job", "position", "role", "jobtitle"]
    for i, h in enumerate(headers):
        if isinstance(h, str) and any(k in h.lower() for k in kws):
            return i
    return None


def normalize_core_title(raw: str) -> str:
    """Strip designations and noise from a raw title to extract a canonical core title."""
    if not raw or not isinstance(raw, str):
        return ""
    t = raw.strip()
    # Remove parenthetical notes first
    t = _PAREN_PATTERN.sub(" ", t)
    # Remove common designation words
    t = _DESIGNATION_PATTERN.sub(" ", t)
    # Remove ordinals / roman numerals
    t = _ORDINAL_PATTERN.sub(" ", t)
    # Remove extra punctuation, excessive whitespace, leading/trailing connectors
    t = re.sub(r'[^\w&/+\- ]', ' ', t)
    t = re.sub(r'\b(of|the|and|for)\b', ' ', t, flags=re.I)  # remove common connectors that may remain
    t = re.sub(r'\s+', ' ', t).strip()
    # Title case for consistency
    return t.title()


def load_json(path: str) -> Dict:
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def write_json(path: str, data: Dict):
    # Overwrite safely: write to tmp then replace
    tmp_path = path + ".tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    os.replace(tmp_path, path)


def read_titles_from_csv(path: str, limit: Optional[int] = None) -> List[str]:
    titles = []
    with open(path, encoding="utf-8", errors="ignore") as f:
        reader = csv.reader(f)
        rows = list(reader)
        if not rows:
            return []
        headers = [str(c) for c in rows[0]]
        title_idx = find_title_column(headers)
        # If not found, treat first column as title
        if title_idx is None:
            title_idx = 0
        for r in rows[1:]:
            if title_idx < len(r):
                t = str(r[title_idx]).strip()
                if t:
                    titles.append(t)
            if limit and len(titles) >= limit:
                break
    return titles


def read_titles_from_xlsx(path: str, limit: Optional[int] = None) -> List[str]:
    if openpyxl is None:
        raise RuntimeError("openpyxl required to read .xlsx files")
    wb = openpyxl.load_workbook(path, read_only=True)
    sheet = wb[wb.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(c) if c is not None else "" for c in rows[0]]
    title_idx = find_title_column(headers)
    if title_idx is None:
        title_idx = 0
    titles = []
    for row in rows[1:]:
        if title_idx < len(row):
            cell = row[title_idx]
            if cell is None:
                continue
            t = str(cell).strip()
            if t:
                titles.append(t)
        if limit and len(titles) >= limit:
            break
    return titles


def read_titles_from_json(path: str, limit: Optional[int] = None) -> List[str]:
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    titles = []
    if isinstance(data, list):
        # list of objects or strings
        if data and isinstance(data[0], dict):
            headers = list(data[0].keys())
            title_key = None
            for k in headers:
                if any(tok in k.lower() for tok in ("title", "job", "role", "position", "jobtitle")):
                    title_key = k
                    break
            if title_key:
                for item in data:
                    t = item.get(title_key)
                    if t:
                        titles.append(str(t).strip())
                        if limit and len(titles) >= limit:
                            break
        else:
            # list of strings
            for item in data:
                if isinstance(item, str):
                    titles.append(item.strip())
                    if limit and len(titles) >= limit:
                        break
    elif isinstance(data, dict):
        # maybe {"rows":[...], "uploaded_columns":[...]} like the frontend payload
        uploaded_columns = data.get("uploaded_columns") or []
        rows = data.get("rows") or []
        if uploaded_columns and rows:
            # reuse normalize_rows behavior is not necessary; assume rows are lists aligned to headers
            title_idx = find_title_column([str(h) for h in uploaded_columns])
            if title_idx is None:
                title_idx = 0
            for r in rows:
                if isinstance(r, (list, tuple)) and title_idx < len(r):
                    v = r[title_idx]
                    if v:
                        titles.append(str(v).strip())
                        if limit and len(titles) >= limit:
                            break
    return titles


def read_titles_from_file(path: str, limit: Optional[int] = None) -> List[str]:
    lower = path.lower()
    if lower.endswith(".csv"):
        return read_titles_from_csv(path, limit)
    if lower.endswith(".xlsx") or lower.endswith(".xlsm") or lower.endswith(".xls"):
        return read_titles_from_xlsx(path, limit)
    if lower.endswith(".json"):
        return read_titles_from_json(path, limit)
    # fallback: try CSV
    return read_titles_from_csv(path, limit)


def fetch_titles_from_db(userid: Optional[str] = None, username: Optional[str] = None, limit: int = 10000) -> List[str]:
    if psycopg2 is None:
        raise RuntimeError("psycopg2 is required for DB access")
    pg_host = os.getenv("PGHOST", "localhost")
    pg_port = int(os.getenv("PGPORT", "5432"))
    pg_user = os.getenv("PGUSER", "postgres")
    pg_password = os.getenv("PGPASSWORD", "") or "orlha"
    pg_db = os.getenv("PGDATABASE", "candidate_db")
    conn = psycopg2.connect(host=pg_host, port=pg_port, user=pg_user, password=pg_password, dbname=pg_db)
    cur = conn.cursor()
    cur.execute("""
        SELECT column_name
        FROM information_schema.columns
        WHERE table_schema='public' AND table_name='process'
    """)
    cols = [r[0].lower() for r in cur.fetchall()]
    preferred_title_col = None
    if 'jobtitle' in cols:
        preferred_title_col = 'jobtitle'
    elif 'role' in cols:
        preferred_title_col = 'role'
    else:
        for c in cols:
            if 'title' in c or 'role' in c:
                preferred_title_col = c
                break
    if not preferred_title_col:
        cur.close(); conn.close()
        raise RuntimeError("No suitable title column found in process table")
    candidate_cols = []
    for c in ['name', 'company', preferred_title_col, 'country', 'linkedinurl', 'sector', 'username', 'userid', 'role_tag']:
        if c in cols and c not in candidate_cols:
            candidate_cols.append(c)
    include_ctid = False
    if 'id' not in cols:
        include_ctid = True
    select_parts = [sql.Identifier(c) for c in candidate_cols]
    if include_ctid:
        select_parts.append(sql.SQL("ctid::text AS _ctid"))
    select_fields = sql.SQL(', ').join(select_parts)
    base_query = sql.SQL("SELECT {fields} FROM process").format(fields=select_fields)
    where_clauses = []
    params = []
    if userid:
        where_clauses.append(sql.SQL("userid = %s")); params.append(str(userid))
    elif username:
        where_clauses.append(sql.SQL("username = %s")); params.append(username)
    if where_clauses:
        where_sql = sql.SQL(" WHERE ") + sql.SQL(" AND ").join(where_clauses)
        query = base_query + where_sql + sql.SQL(" ORDER BY name NULLS LAST LIMIT %s")
        params.append(limit)
    else:
        query = base_query + sql.SQL(" ORDER BY name NULLS LAST LIMIT %s")
        params = [limit]
    cur.execute(query.as_string(conn), params)
    fetched = cur.fetchall()
    titles = []
    title_idx = candidate_cols.index(preferred_title_col)
    for r in fetched:
        if title_idx < len(r):
            t = r[title_idx]
            if t:
                titles.append(str(t).strip())
    cur.close(); conn.close()
    return titles


def heuristic_family_for_title(title: str) -> str:
    """Fallback family heuristics if map_titles_to_families... not available or returns empty."""
    t = title.lower()
    if any(k in t for k in _PROGRAMMING_KEYWORDS):
        return "Programming"
    if any(k in t for k in _ART_KEYWORDS):
        # If animation-specific, prefer Animation
        if 'anim' in t or 'animation' in t:
            return "Animation"
        return "Art"
    if any(k in t for k in _DESIGN_KEYWORDS):
        return "Design"
    if any(k in t for k in _QA_KEYWORDS):
        return "QA"
    if any(k in t for k in _AUDIO_KEYWORDS):
        return "Audio"
    if any(k in t for k in _PROD_KEYWORDS):
        return "Production"
    if any(k in t for k in _DATA_KEYWORDS):
        return "Data Science"
    if any(k in t for k in _FINANCE_KEYWORDS):
        return "Finance"
    # Default fallback
    return "Corporate"


def flatten_existing_titles(jobfamilyroles: Dict[str, List[str]]) -> Set[str]:
    s = set()
    for fam, titles in jobfamilyroles.items():
        for t in titles:
            if isinstance(t, str) and t.strip():
                s.add(t.strip().lower())
    return s


def format_additions_for_log(additions: List[Tuple[str, str]], max_items: int = 20) -> str:
    """
    Return a safe, truncated string describing additions for logging.
    Examples:
      "Software Engineer, Software Engineering"
      "A, B, C, ... (+5 more)"
    """
    if not additions:
        return ""
    titles = [t for _, t in additions]
    if len(titles) <= max_items:
        return ", ".join(titles)
    else:
        shown = titles[:max_items]
        more = len(titles) - max_items
        return f"{', '.join(shown)} (+{more} more)"


def main(argv: List[str]):
    ap = argparse.ArgumentParser(description="Update JobFamilyRoles in static/data_sorter.json from process table or file.")
    ap.add_argument("--file", "-f", help="Path to local file (CSV, XLSX, JSON) containing titles", default=None)
    ap.add_argument("--db", action="store_true", help="Fetch titles from Postgres 'process' table (uses PG env vars).")
    ap.add_argument("--userid", help="If using --db, restrict to this userid", default=None)
    ap.add_argument("--username", help="If using --db, restrict to this username", default=None)
    ap.add_argument("--limit", type=int, help="Max number of rows to fetch from DB/file", default=5000)
    ap.add_argument("--json", help=f"Path to data_sorter.json (default: {DEFAULT_JSON_PATH})", default=DEFAULT_JSON_PATH)
    ap.add_argument("--commit", action="store_true", help="Write changes back to data_sorter.json (overwrite). If omitted, runs a dry-run.")
    ap.add_argument("--preview-only", action="store_true", help="Alias for dry-run (do not write).")
    args = ap.parse_args(argv)

    json_path = args.json
    if not os.path.exists(json_path):
        print(f"ERROR: data_sorter.json not found at {json_path}", file=sys.stderr)
        sys.exit(2)

    try:
        cfg = load_json(json_path)
    except Exception as e:
        print(f"ERROR: failed to load {json_path}: {e}", file=sys.stderr)
        sys.exit(2)

    jobfamilyroles = cfg.get("JobFamilyRoles") or {}
    existing_titles_set = flatten_existing_titles(jobfamilyroles)

    # 1) Read titles from source
    titles_raw: List[str] = []
    try:
        if args.db:
            print("Fetching titles from DB...")
            titles_raw = fetch_titles_from_db(userid=args.userid, username=args.username, limit=args.limit)
        elif args.file:
            print(f"Reading titles from file {args.file} ...")
            titles_raw = read_titles_from_file(args.file, limit=args.limit)
        else:
            print("No source specified. Use --db or --file. Exiting.", file=sys.stderr)
            sys.exit(1)
    except Exception as e:
        print(f"ERROR: failed to read titles: {e}", file=sys.stderr)
        sys.exit(2)

    if not titles_raw:
        print("No titles found from source. Nothing to do.")
        sys.exit(0)

    # 2) Normalize to core titles
    core_titles_map: Dict[str, str] = {}  # original -> core
    for t in titles_raw:
        core = normalize_core_title(t)
        if core:
            core_titles_map[t] = core

    unique_cores = list({c for c in core_titles_map.values() if c})
    # Remove entries that already exist in JSON (case-insensitive)
    to_consider = [c for c in unique_cores if c.strip().lower() not in existing_titles_set]
    if not to_consider:
        print("No new core titles to add (after stripping designations).")
        sys.exit(0)

    print(f"Found {len(unique_cores)} unique core titles; {len(to_consider)} appear new vs data_sorter.json")

    # 3) Use title2vec_service to map titles to families (if available)
    predicted_families: Dict[str, str] = {}
    if map_titles_to_families_and_seniority is not None:
        try:
            print("Running family classifier (title2vec_service)...")
            fams, _ = map_titles_to_families_and_seniority(to_consider)
            for core, fam in zip(to_consider, fams):
                predicted_families[core] = fam or ""
        except Exception as e:
            print(f"Warning: classifier failed: {e}; will use heuristics.", file=sys.stderr)

    # 4) For each new core, decide family and append
    additions: List[Tuple[str, str]] = []  # (family, title)
    for core in to_consider:
        fam = predicted_families.get(core, "")
        if not fam:
            fam = heuristic_family_for_title(core)
        fam = fam or "Corporate"
        # Ensure family key exists
        if fam not in jobfamilyroles:
            # create family if it doesn't exist
            jobfamilyroles[fam] = []
        # Avoid duplicate (case-insensitive)
        existing_lower = {t.lower() for t in jobfamilyroles.get(fam, []) if isinstance(t, str)}
        if core.lower() not in existing_lower:
            # Final validation step: use JobBERT local classifier (option B) and only call Gemini for ambiguous cases.
            accept = None
            try:
                score = jobbert_score_local(core, model_id=os.getenv("VC_JOBBERT_MODEL", JOBBERT_MODEL_ENV))
            except Exception:
                score = None

            if score is None:
                # Classifier unavailable or errored
                if JOBBERT_FALLBACK_MODE == "permissive":
                    accept = True
                else:
                    # conservative: do not accept if no classifier result; attempt Gemini if available
                    if _HAS_GEMINI and gemini_context is not None:
                        try:
                            accept = is_likely_job_title_gemini(core)
                        except Exception:
                            accept = False
                    else:
                        accept = False
            else:
                # Decide based on thresholds
                if score >= JOBBERT_HIGH:
                    accept = True
                elif score <= JOBBERT_LOW:
                    accept = False
                else:
                    # Ambiguous -> consult Gemini if available
                    if _HAS_GEMINI and gemini_context is not None:
                        try:
                            accept = is_likely_job_title_gemini(core)
                        except Exception:
                            accept = False if JOBBERT_FALLBACK_MODE == "conservative" else True
                    else:
                        accept = False if JOBBERT_FALLBACK_MODE == "conservative" else True

            if not accept:
                print(f"Skipping addition (JobBERT/Gemini validation): {core} -> {fam}")
                continue

            jobfamilyroles[fam].append(core)
            additions.append((fam, core))

    # 5) Summary and optional write
    if not additions:
        print("No additions performed (duplicates or no suitable family).")
        sys.exit(0)

    print("Planned additions:")
    for fam, title in additions:
        print(f"  - {title} -> {fam}")

    if args.commit and not args.preview_only:
        # update cfg and write back
        try:
            # Add RecentUpdates entries so frontend can display server-origin changes immediately
            recent = cfg.get("RecentUpdates", [])
            for fam, title in additions:
                try:
                    recent_entry = {
                        "ts": int(time.time()),
                        "iso": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
                        "action": "Added",
                        "family": fam,
                        "title": title
                    }
                    recent.insert(0, recent_entry)
                except Exception:
                    # non-fatal; continue adding others
                    pass
            # keep only the most recent 100 entries
            cfg["RecentUpdates"] = recent[:100]
        except Exception as e:
            print(f"Warning: failed to append RecentUpdates: {e}", file=sys.stderr)

        cfg["JobFamilyRoles"] = jobfamilyroles
        try:
            write_json(json_path, cfg)
            # friendly console output
            print(f"Successfully updated {json_path} (overwrote). Added {len(additions)} titles.")
            # structured log for server logs including a short list of titles
            titles_str = format_additions_for_log(additions, max_items=20)
            logger.info("Updated %s with %d new titles: %s", json_path, len(additions), titles_str)
        except Exception as e:
            print(f"ERROR: failed to write {json_path}: {e}", file=sys.stderr)
            logger.exception("Failed to write updated JSON %s", json_path)
            sys.exit(3)
    else:
        print("\nDry-run mode (no file overwritten). Use --commit to apply changes.")
        # Print suggested JSON fragment for quick copy/paste
        suggestion = { "JobFamilyRoles": { k: v for k, v in jobfamilyroles.items() if any(t in [x for _, x in additions] for t in v) or k in {fam for fam, _ in additions} } }
        print("\nSuggested additions JSON fragment (for reference):")
        print(json.dumps(suggestion, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main(sys.argv[1:])